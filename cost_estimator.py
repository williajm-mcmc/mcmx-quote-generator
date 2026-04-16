# cost_estimator.py
"""
Cost Estimator tab — replicates the OC Lev 4-ASPCostReport Excel workflow.

Sections:
  1. Consulting Effort  — hours × hourly rate per labour category
  2. Third Party        — free-form description / cost rows
  3. Hardware / Software / Materials — qty × unit cost rows
  4. Travel & Expenses  — airfare / hotel / food / car
  5. Summary            — rollup → risk → margin → resale → profit
"""

import math
import json
import os
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QScrollArea, QFrame, QSizePolicy, QDoubleSpinBox, QAbstractItemView,
    QFileDialog, QMessageBox,
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QColor, QFont, QDoubleValidator, QIntValidator

# ── Shared colour constants ──────────────────────────────────────────────────
_RED      = "#920d2e"
_RED_DARK = "#7a0b27"
_TEXT     = "#1a0509"
_SUBTEXT  = "#3a3a5c"
_BG       = "#ffffff"
_BG_CARD  = "#ffffff"
_BG_HDR   = "#f9f0f2"
_BORDER   = "#d6c0c5"
_BLUE_BG  = "#e8f0fe"
_BLUE_FG  = "#1a3a6e"
_GRAY_BG  = "#f4f6fa"

_FIELD_STYLE = (
    "QLineEdit { color:#1a0509; background:#ffffff;"
    "  border:1px solid #d6c0c5; border-radius:4px;"
    "  padding:4px 8px; font-size:12px; }"
    "QLineEdit:focus { border-color:#920d2e; }"
)
_CALC_STYLE = (
    "QLineEdit { color:#1a3a6e; background:#e8f0fe;"
    "  border:1px solid #b8cce4; border-radius:4px;"
    "  padding:4px 8px; font-size:12px; font-weight:600; }"
)
_BTN_STYLE = (
    "QPushButton { background:#f4f6fa; color:#3a3a5c;"
    "  border:1px solid #dde1e7; border-radius:4px;"
    "  padding:4px 12px; font-size:11px; }"
    "QPushButton:hover { background:#e8ecf5; border-color:#b0b8d0; }"
    "QPushButton:pressed { background:#d8def0; }"
)
_DEL_BTN_STYLE = (
    "QPushButton { background:transparent; color:#e05252;"
    "  border:1px solid #e05252; border-radius:4px;"
    "  padding:4px 12px; font-size:11px; }"
    "QPushButton:hover { background:#fdf2f2; }"
)
_TABLE_STYLE = (
    "QTableWidget { background:#ffffff; border:1px solid #d6c0c5;"
    "  border-radius:4px; gridline-color:#e8dde0; font-size:11px; }"
    "QHeaderView::section { background:#f4f6fa; color:#3a3a5c;"
    "  font-size:11px; font-weight:600; border:none;"
    "  border-bottom:1px solid #d6c0c5; padding:4px 6px; }"
    "QTableWidget::item { padding:3px 6px; }"
    "QTableWidget::item:selected { background:#f5d0da; color:#920d2e; }"
)


# ── Helper: read a numeric QLineEdit ────────────────────────────────────────
def _num(widget: QLineEdit, default=0.0) -> float:
    try:
        return float(widget.text().replace(",", "").replace("$", "").strip() or default)
    except ValueError:
        return default


def _fmt_dollar(v: float) -> str:
    return f"${math.ceil(v):,}"


def _fmt_plain(v: float) -> str:
    if v == int(v):
        return f"{int(v):,}"
    return f"{v:,.2f}"


# ── Card widget (collapsible) ────────────────────────────────────────────────
class _Card(QFrame):
    """Collapsible section card matching the app's visual language."""

    def __init__(self, title: str, parent=None):
        super().__init__(parent)
        self.setStyleSheet(
            f"QFrame {{ background:{_BG_CARD}; border:1px solid {_BORDER};"
            "  border-radius:8px; }"
        )
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Header bar
        hdr = QWidget()
        hdr.setStyleSheet(
            f"QWidget {{ background:{_BG_HDR}; border-radius:7px 7px 0 0;"
            f"  border-bottom:1px solid {_BORDER}; }}"
        )
        hdr.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        hdr.setFixedHeight(40)
        hdr_row = QHBoxLayout(hdr)
        hdr_row.setContentsMargins(12, 0, 8, 0)
        hdr_row.setSpacing(8)
        self._hdr_row = hdr_row  # expose for subclasses

        self._toggle_btn = QPushButton("▼")
        self._toggle_btn.setFixedSize(22, 22)
        self._toggle_btn.setStyleSheet(
            f"QPushButton {{ background:transparent; border:none;"
            f"  color:{_RED}; font-size:12px; font-weight:700; padding:0; }}"
        )
        self._toggle_btn.clicked.connect(self._toggle)
        hdr_row.addWidget(self._toggle_btn)

        self._title_lbl = QLabel(title)
        self._title_lbl.setStyleSheet(
            f"QLabel {{ font-size:13px; font-weight:700; color:{_RED};"
            "  background:transparent; border:none; padding:0; }"
        )
        hdr_row.addWidget(self._title_lbl)
        hdr_row.addStretch()

        # slot for subtotal badge in header
        self._subtotal_lbl = QLabel("")
        self._subtotal_lbl.setStyleSheet(
            f"QLabel {{ font-size:12px; font-weight:700; color:{_RED};"
            "  background:transparent; border:none; padding:0 4px; }"
        )
        hdr_row.addWidget(self._subtotal_lbl)

        root.addWidget(hdr)

        # Body
        self._body = QWidget()
        self._body.setStyleSheet(
            "QWidget { background:transparent; border:none; }"
        )
        self._body_layout = QVBoxLayout(self._body)
        self._body_layout.setContentsMargins(14, 12, 14, 14)
        self._body_layout.setSpacing(8)
        root.addWidget(self._body)

        self._collapsed = False

    def _toggle(self):
        self._collapsed = not self._collapsed
        self._body.setVisible(not self._collapsed)
        self._toggle_btn.setText("▶" if self._collapsed else "▼")

    def set_subtotal(self, text: str):
        self._subtotal_lbl.setText(text)


# ── Numeric-only QLineEdit ───────────────────────────────────────────────────
class _NumEdit(QLineEdit):
    """QLineEdit that accepts decimal numbers and emits on change."""

    def __init__(self, placeholder="0", decimals=2, parent=None):
        super().__init__(parent)
        self.setPlaceholderText(placeholder)
        self.setStyleSheet(_FIELD_STYLE)
        self.setAlignment(Qt.AlignmentFlag.AlignRight)
        self._decimals = decimals

    def value(self) -> float:
        return _num(self)


# ── Calculated (read-only) display field ────────────────────────────────────
class _CalcEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
        self.setStyleSheet(_CALC_STYLE)
        self.setAlignment(Qt.AlignmentFlag.AlignRight)


# ══════════════════════════════════════════════════════════════════════════════
# Section 1 — Consulting Effort
# ══════════════════════════════════════════════════════════════════════════════
class _ConsultingCard(_Card):
    _ROWS = ["On-Site Time (hrs)", "Travel Time (hrs)", "Other (hrs)"]

    def __init__(self, on_change, parent=None):
        super().__init__("2.  Consulting Effort", parent)
        self._on_change = on_change

        # Hourly rate row
        rate_row = QHBoxLayout()
        rate_row.setSpacing(8)
        rate_row.addWidget(QLabel("Hourly Rate  ($):"))
        self._rate = _NumEdit("47.00")
        self._rate.setText("47.00")
        self._rate.setFixedWidth(120)
        self._rate.textChanged.connect(self._recalc)
        rate_row.addWidget(self._rate)
        rate_row.addStretch()
        self._body_layout.addLayout(rate_row)

        # Table: Labour category | Hours | Rate | Cost
        self._table = QTableWidget(len(self._ROWS), 4)
        self._table.setStyleSheet(_TABLE_STYLE)
        self._table.setHorizontalHeaderLabels(["Category", "Hours", "Rate ($/hr)", "Cost"])
        self._table.verticalHeader().setVisible(False)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        hh = self._table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for c in (1, 2, 3):
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.Fixed)
        self._table.setColumnWidth(1, 90)
        self._table.setColumnWidth(2, 110)
        self._table.setColumnWidth(3, 110)
        self._table.setFixedHeight(32 * len(self._ROWS) + 34)

        self._hour_edits = []
        self._rate_edits = []
        self._cost_edits = []
        for r, lbl in enumerate(self._ROWS):
            # Category label
            cat = QTableWidgetItem(lbl)
            cat.setFlags(Qt.ItemFlag.ItemIsEnabled)
            cat.setForeground(QColor(_SUBTEXT))
            self._table.setItem(r, 0, cat)
            # Hours
            hrs = _NumEdit("0")
            hrs.textChanged.connect(self._recalc)
            self._table.setCellWidget(r, 1, hrs)
            self._hour_edits.append(hrs)
            # Rate — editable; empty → uses global rate
            rate_cell = _NumEdit()
            rate_cell.setPlaceholderText("= global")
            rate_cell.setToolTip(
                "Leave blank to use the global hourly rate above.\n"
                "Enter a value to override for this row only."
            )
            rate_cell.textChanged.connect(self._recalc)
            self._table.setCellWidget(r, 2, rate_cell)
            self._rate_edits.append(rate_cell)
            # Cost (calculated)
            cost_cell = _CalcEdit()
            self._table.setCellWidget(r, 3, cost_cell)
            self._cost_edits.append(cost_cell)

        self._body_layout.addWidget(self._table)

        # Subtotal row
        sub_row = QHBoxLayout()
        sub_row.addStretch()
        sub_row.addWidget(QLabel("Subtotal:"))
        self._subtotal_edit = _CalcEdit()
        self._subtotal_edit.setFixedWidth(130)
        sub_row.addWidget(self._subtotal_edit)
        self._body_layout.addLayout(sub_row)

        self._recalc()

    def _recalc(self):
        global_rate = _num(self._rate)
        total = 0.0
        for r in range(len(self._ROWS)):
            hrs = self._hour_edits[r].value()
            # Use per-row rate if filled in, otherwise fall back to global
            row_rate_text = self._rate_edits[r].text().strip()
            row_rate = _num(self._rate_edits[r]) if row_rate_text else global_rate
            cost = hrs * row_rate
            total += cost
            self._cost_edits[r].setText(_fmt_dollar(cost))
        self._subtotal_edit.setText(_fmt_dollar(total))
        self.set_subtotal(_fmt_dollar(total))
        self._on_change()

    def subtotal(self) -> float:
        global_rate = _num(self._rate)
        total = 0.0
        for r, ed in enumerate(self._hour_edits):
            row_rate_text = self._rate_edits[r].text().strip()
            row_rate = _num(self._rate_edits[r]) if row_rate_text else global_rate
            total += ed.value() * row_rate
        return total

    def get_data(self) -> dict:
        return {
            "rate":       self._rate.text(),
            "rows":       [ed.text() for ed in self._hour_edits],
            "row_rates":  [ed.text() for ed in self._rate_edits],
        }

    def restore_data(self, d: dict):
        self._rate.setText(d.get("rate", ""))
        for i, val in enumerate(d.get("rows", [])):
            if i < len(self._hour_edits):
                self._hour_edits[i].setText(val)
        for i, val in enumerate(d.get("row_rates", [])):
            if i < len(self._rate_edits):
                self._rate_edits[i].setText(val)


# ══════════════════════════════════════════════════════════════════════════════
# Section 2 — Third Party
# ══════════════════════════════════════════════════════════════════════════════
class _ThirdPartyCard(_Card):
    def __init__(self, on_change, parent=None):
        super().__init__("3.  Third Party", parent)
        self._on_change = on_change
        self._rows: list[tuple[QLineEdit, _NumEdit, _CalcEdit]] = []

        # Table
        self._table = QTableWidget(0, 3)
        self._table.setStyleSheet(_TABLE_STYLE)
        self._table.setHorizontalHeaderLabels(["Description", "Qty", "Cost Total"])
        self._table.verticalHeader().setVisible(False)
        hh = self._table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self._table.setColumnWidth(1, 80)
        self._table.setColumnWidth(2, 130)
        self._body_layout.addWidget(self._table)

        # Add / Remove buttons
        btn_row = QHBoxLayout()
        add_btn = QPushButton("＋  Add Row")
        add_btn.setStyleSheet(_BTN_STYLE)
        add_btn.clicked.connect(self.add_row)
        del_btn = QPushButton("－  Remove Selected")
        del_btn.setStyleSheet(_DEL_BTN_STYLE)
        del_btn.clicked.connect(self._remove_row)
        btn_row.addWidget(add_btn); btn_row.addWidget(del_btn); btn_row.addStretch()
        self._body_layout.addLayout(btn_row)

        # Subtotal
        sub_row = QHBoxLayout()
        sub_row.addStretch()
        sub_row.addWidget(QLabel("Subtotal:"))
        self._subtotal_edit = _CalcEdit()
        self._subtotal_edit.setFixedWidth(130)
        sub_row.addWidget(self._subtotal_edit)
        self._body_layout.addLayout(sub_row)

        self.add_row()

    def add_row(self):
        r = self._table.rowCount()
        self._table.insertRow(r)
        self._table.setRowHeight(r, 36)

        desc = QLineEdit()
        desc.setStyleSheet(_FIELD_STYLE)
        self._table.setCellWidget(r, 0, desc)

        qty = _NumEdit("0", decimals=0)
        qty.setFixedWidth(70)
        qty.textChanged.connect(self._recalc)
        self._table.setCellWidget(r, 1, qty)

        cost = _NumEdit("0.00")
        cost.textChanged.connect(self._recalc)
        self._table.setCellWidget(r, 2, cost)

        self._rows.append((desc, qty, cost))
        self._update_height()
        self._recalc()

    def _remove_row(self):
        rows = sorted({i.row() for i in self._table.selectedIndexes()}, reverse=True)
        for r in rows:
            self._table.removeRow(r)
            if r < len(self._rows):
                self._rows.pop(r)
        self._update_height()
        self._recalc()

    def _update_height(self):
        n = max(1, self._table.rowCount())
        self._table.setFixedHeight(36 * n + 38)

    def _recalc(self):
        total = 0.0
        for _, qty, cost in self._rows:
            total += qty.value() * cost.value()
        self._subtotal_edit.setText(_fmt_dollar(total))
        self.set_subtotal(_fmt_dollar(total))
        self._on_change()

    def subtotal(self) -> float:
        return sum(q.value() * c.value() for _, q, c in self._rows)

    def get_data(self) -> dict:
        return {"rows": [(d.text(), q.text(), c.text()) for d, q, c in self._rows]}

    def restore_data(self, d: dict):
        while self._table.rowCount():
            self._table.removeRow(0)
        self._rows.clear()
        for desc, qty, cost in d.get("rows", []):
            self.add_row()
            self._rows[-1][0].setText(desc)
            self._rows[-1][1].setText(qty)
            self._rows[-1][2].setText(cost)


# ══════════════════════════════════════════════════════════════════════════════
# Section 3 — Hardware / Software / Materials
# ══════════════════════════════════════════════════════════════════════════════
class _MaterialsCard(_Card):
    def __init__(self, on_change, parent=None):
        super().__init__("4.  Hardware / Software / Materials", parent)
        self._on_change = on_change
        self._rows: list[tuple[QLineEdit, _NumEdit, _NumEdit, _CalcEdit]] = []

        self._table = QTableWidget(0, 4)
        self._table.setStyleSheet(_TABLE_STYLE)
        self._table.setHorizontalHeaderLabels(["Description", "Qty", "Cost Per", "Cost Total"])
        self._table.verticalHeader().setVisible(False)
        hh = self._table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for c in (1, 2, 3):
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.Fixed)
        self._table.setColumnWidth(1, 70)
        self._table.setColumnWidth(2, 110)
        self._table.setColumnWidth(3, 120)
        self._body_layout.addWidget(self._table)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("＋  Add Row"); add_btn.setStyleSheet(_BTN_STYLE)
        add_btn.clicked.connect(self.add_row)
        del_btn = QPushButton("－  Remove Selected"); del_btn.setStyleSheet(_DEL_BTN_STYLE)
        del_btn.clicked.connect(self._remove_row)
        btn_row.addWidget(add_btn); btn_row.addWidget(del_btn); btn_row.addStretch()
        self._body_layout.addLayout(btn_row)

        sub_row = QHBoxLayout()
        sub_row.addStretch()
        sub_row.addWidget(QLabel("Subtotal:"))
        self._subtotal_edit = _CalcEdit()
        self._subtotal_edit.setFixedWidth(130)
        sub_row.addWidget(self._subtotal_edit)
        self._body_layout.addLayout(sub_row)

        self.add_row()

    def add_row(self):
        r = self._table.rowCount()
        self._table.insertRow(r)
        self._table.setRowHeight(r, 36)

        desc = QLineEdit(); desc.setStyleSheet(_FIELD_STYLE)
        self._table.setCellWidget(r, 0, desc)

        qty = _NumEdit("0", decimals=0)
        qty.textChanged.connect(self._recalc)
        self._table.setCellWidget(r, 1, qty)

        cost_per = _NumEdit("0.00")
        cost_per.textChanged.connect(self._recalc)
        self._table.setCellWidget(r, 2, cost_per)

        total = _CalcEdit()
        self._table.setCellWidget(r, 3, total)

        self._rows.append((desc, qty, cost_per, total))
        self._update_height()
        self._recalc()

    def _remove_row(self):
        rows = sorted({i.row() for i in self._table.selectedIndexes()}, reverse=True)
        for r in rows:
            self._table.removeRow(r)
            if r < len(self._rows):
                self._rows.pop(r)
        self._update_height()
        self._recalc()

    def _update_height(self):
        n = max(1, self._table.rowCount())
        self._table.setFixedHeight(36 * n + 38)

    def _recalc(self):
        total = 0.0
        for _, qty, cost_per, cost_total in self._rows:
            v = qty.value() * cost_per.value()
            total += v
            cost_total.setText(_fmt_dollar(v))
        self._subtotal_edit.setText(_fmt_dollar(total))
        self.set_subtotal(_fmt_dollar(total))
        self._on_change()

    def subtotal(self) -> float:
        return sum(q.value() * c.value() for _, q, c, _ in self._rows)

    def get_data(self) -> dict:
        return {"rows": [(d.text(), q.text(), c.text()) for d, q, c, _ in self._rows]}

    def restore_data(self, d: dict):
        while self._table.rowCount():
            self._table.removeRow(0)
        self._rows.clear()
        for desc, qty, cost in d.get("rows", []):
            self.add_row()
            self._rows[-1][0].setText(desc)
            self._rows[-1][1].setText(qty)
            self._rows[-1][2].setText(cost)


# ══════════════════════════════════════════════════════════════════════════════
# Section 4 — Travel & Expenses
# ══════════════════════════════════════════════════════════════════════════════
class _TravelCard(_Card):
    _LABELS   = ["Airfare",   "Hotel (nights)", "Food (meals)", "Car (miles)"]
    _QTY_HINT = ["0",         "0",              "0",            "0"          ]
    _RATE_HINT= ["0.00",      "0.00",           "0.00",         "0.670"      ]
    _QTY_LBL  = ["Qty",       "Nights",         "Meals",        "Miles"      ]
    _RATE_LBL = ["Cost Each", "$/Night",        "$/Meal",       "$/Mile"     ]

    def __init__(self, on_change, parent=None):
        super().__init__("5.  Travel & Expenses", parent)
        self._on_change = on_change

        self._table = QTableWidget(len(self._LABELS), 4)
        self._table.setStyleSheet(_TABLE_STYLE)
        self._table.setHorizontalHeaderLabels(["Category", "Qty", "Rate", "Cost"])
        self._table.verticalHeader().setVisible(False)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        hh = self._table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for c in (1, 2, 3):
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.Fixed)
        self._table.setColumnWidth(1, 90)
        self._table.setColumnWidth(2, 110)
        self._table.setColumnWidth(3, 110)
        self._table.setFixedHeight(32 * len(self._LABELS) + 34)

        self._qty_edits  = []
        self._rate_edits = []
        self._cost_edits = []
        for r, lbl in enumerate(self._LABELS):
            cat = QTableWidgetItem(lbl)
            cat.setFlags(Qt.ItemFlag.ItemIsEnabled)
            cat.setForeground(QColor(_SUBTEXT))
            self._table.setItem(r, 0, cat)

            qty = _NumEdit(self._QTY_HINT[r])
            qty.setText("0")
            qty.setToolTip(self._QTY_LBL[r])
            qty.textChanged.connect(self._recalc)
            self._table.setCellWidget(r, 1, qty)
            self._qty_edits.append(qty)

            rate = _NumEdit(self._RATE_HINT[r])
            rate.setToolTip(self._RATE_LBL[r])
            rate.textChanged.connect(self._recalc)
            self._table.setCellWidget(r, 2, rate)
            self._rate_edits.append(rate)

            cost = _CalcEdit()
            self._table.setCellWidget(r, 3, cost)
            self._cost_edits.append(cost)

        self._body_layout.addWidget(self._table)

        sub_row = QHBoxLayout()
        sub_row.addStretch()
        sub_row.addWidget(QLabel("Subtotal:"))
        self._subtotal_edit = _CalcEdit()
        self._subtotal_edit.setFixedWidth(130)
        sub_row.addWidget(self._subtotal_edit)
        self._body_layout.addLayout(sub_row)

        self._recalc()

    def _recalc(self):
        total = 0.0
        for r in range(len(self._LABELS)):
            v = self._qty_edits[r].value() * self._rate_edits[r].value()
            total += v
            self._cost_edits[r].setText(_fmt_dollar(v))
        self._subtotal_edit.setText(_fmt_dollar(total))
        self.set_subtotal(_fmt_dollar(total))
        self._on_change()

    def subtotal(self) -> float:
        return sum(q.value() * r.value()
                   for q, r in zip(self._qty_edits, self._rate_edits))

    def get_data(self) -> dict:
        return {
            "qty":  [e.text() for e in self._qty_edits],
            "rate": [e.text() for e in self._rate_edits],
        }

    def restore_data(self, d: dict):
        for i, v in enumerate(d.get("qty", [])):
            if i < len(self._qty_edits): self._qty_edits[i].setText(v)
        for i, v in enumerate(d.get("rate", [])):
            if i < len(self._rate_edits): self._rate_edits[i].setText(v)


# ══════════════════════════════════════════════════════════════════════════════
# Section 5 — Summary
# ══════════════════════════════════════════════════════════════════════════════
class _SummaryCard(QFrame):
    """Always-visible summary / rollup panel at the bottom."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(
            f"QFrame {{ background:{_BG_CARD}; border:2px solid {_RED};"
            "  border-radius:8px; }"
        )

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 16)
        layout.setSpacing(8)

        title = QLabel("Cost Summary")
        title.setStyleSheet(
            f"QLabel {{ font-size:14px; font-weight:800; color:{_RED};"
            "  border:none; padding:0; }"
        )
        layout.addWidget(title)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"border:none; border-top:1px solid {_BORDER};")
        layout.addWidget(sep)

        grid = QWidget()
        grid.setStyleSheet("QWidget { background:transparent; border:none; }")
        gl = QVBoxLayout(grid)
        gl.setContentsMargins(0, 0, 0, 0)
        gl.setSpacing(6)

        def _row(label: str, is_input=False, hint="0.00", bold=False, accent=False):
            row = QHBoxLayout()
            lbl = QLabel(label)
            lbl.setFixedWidth(240)
            col = _RED if accent else _SUBTEXT
            w = 700 if bold else 400
            lbl.setStyleSheet(
                f"QLabel {{ font-size:12px; font-weight:{w}; color:{col};"
                "  border:none; padding:0; }"
            )
            row.addWidget(lbl)
            if is_input:
                ed = _NumEdit(hint)
                ed.setFixedWidth(130)
                val = ed
            else:
                ed = _CalcEdit()
                ed.setFixedWidth(130)
                val = ed
            row.addWidget(val)
            row.addStretch()
            gl.addLayout(row)
            return val

        self._cost_before_risk = _row("Cost Before Risk:", bold=True)
        self._risk_pct         = _row("Risk / Insurance  (%):", is_input=True, hint="0.00")
        self._risk_cost        = _row("  Risk Cost:")
        self._estimated_cost   = _row("Estimated Cost:", bold=True)
        self._margin_pct       = _row("Margin  (%):", is_input=True, hint="25.0")
        self._resale           = _row("Recommended Resale:", bold=True, accent=True)
        self._profit           = _row("Profit:", bold=True, accent=True)

        layout.addWidget(grid)

    def update(self, consulting: float, third_party: float,
               materials: float, travel: float):
        cost_before = consulting + third_party + materials + travel
        risk_pct    = _num(self._risk_pct)
        risk_cost   = cost_before * risk_pct / 100.0
        estimated   = cost_before + risk_cost
        margin_pct  = _num(self._margin_pct)
        if 0 < margin_pct < 100:
            resale = estimated / (1.0 - margin_pct / 100.0)
        else:
            resale = estimated
        profit = resale - estimated

        self._cost_before_risk.setText(_fmt_dollar(cost_before))
        self._risk_cost.setText(_fmt_dollar(risk_cost))
        self._estimated_cost.setText(_fmt_dollar(estimated))
        self._resale.setText(_fmt_dollar(resale))
        self._profit.setText(_fmt_dollar(profit))

    def get_data(self) -> dict:
        return {
            "risk_pct":   self._risk_pct.text(),
            "margin_pct": self._margin_pct.text(),
        }

    def restore_data(self, d: dict):
        self._risk_pct.setText(d.get("risk_pct", ""))
        self._margin_pct.setText(d.get("margin_pct", "25.0"))


# ══════════════════════════════════════════════════════════════════════════════
# Top-level Cost Estimator widget
# ══════════════════════════════════════════════════════════════════════════════
class CostEstimatorWidget(QWidget):
    """
    Drop-in widget for the Cost Estimator tab.
    Mirrors the Excel OC Lev 4-ASPCostReport structure.
    """

    def __init__(self, parent=None):
        super().__init__(parent)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Scroll area for the cards
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet(
            "QScrollArea { background:#f0f2f7; border:none; }"
            "QScrollBar:vertical { background:#f0f2f7; width:8px; margin:0; }"
            "QScrollBar::handle:vertical { background:#c8cedd; border-radius:4px; min-height:20px; }"
            "QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0; }"
        )

        content = QWidget()
        content.setStyleSheet("QWidget { background:#f0f2f7; }")
        cl = QVBoxLayout(content)
        cl.setContentsMargins(20, 16, 20, 16)
        cl.setSpacing(12)

        # Project info strip
        info_card = _Card("Project Info", parent=content)
        info_row = QHBoxLayout()
        info_row.setSpacing(12)
        for lbl_txt, attr, hint in [
            ("Client:",   "_client",  "Client name…"),
            ("Project:",  "_proj",    "Project name…"),
            ("Date:",     "_date",    "MM/DD/YYYY"),
        ]:
            lbl = QLabel(lbl_txt)
            lbl.setStyleSheet("QLabel { border:none; }")
            info_row.addWidget(lbl)
            ed = QLineEdit()
            ed.setPlaceholderText(hint)
            ed.setStyleSheet(_FIELD_STYLE)
            setattr(self, attr, ed)
            info_row.addWidget(ed)
        info_card._body_layout.addLayout(info_row)
        cl.addWidget(info_card)

        # Cost sections: Consulting, Third Party, Materials, Travel
        self._consulting  = _ConsultingCard(self._recalculate, parent=content)
        self._third_party = _ThirdPartyCard(self._recalculate, parent=content)
        self._materials   = _MaterialsCard(self._recalculate, parent=content)
        self._travel      = _TravelCard(self._recalculate, parent=content)
        cl.addWidget(self._consulting)
        cl.addWidget(self._third_party)
        cl.addWidget(self._materials)
        cl.addWidget(self._travel)

        # Summary (always visible, not collapsible)
        self._summary = _SummaryCard(parent=content)
        # Wire summary inputs to recalculate
        self._summary._risk_pct.textChanged.connect(self._recalculate)
        self._summary._margin_pct.textChanged.connect(self._recalculate)
        cl.addWidget(self._summary)
        cl.addStretch()

        scroll.setWidget(content)
        root.addWidget(scroll)

        # ── Bottom action bar ──────────────────────────────────────────────
        bar = QWidget()
        bar.setStyleSheet(
            "QWidget { background:#f9f0f2; border-top:1px solid #d6c0c5; }"
        )
        bar.setFixedHeight(52)
        bar_row = QHBoxLayout(bar)
        bar_row.setContentsMargins(16, 8, 16, 8)
        bar_row.setSpacing(10)

        bar_row.addStretch()

        load_btn = QPushButton("⬆  Import .mcmxc…")
        load_btn.setStyleSheet(
            "QPushButton { background:#ffffff; color:#920d2e;"
            "  border:1px solid #d6c0c5; border-radius:5px;"
            "  padding:7px 18px; font-size:12px; font-weight:600; }"
            "QPushButton:hover { background:#fdf0f3; border-color:#920d2e; }"
        )
        load_btn.clicked.connect(self.import_cost_sheet)
        bar_row.addWidget(load_btn)

        save_btn = QPushButton("💾  Save .mcmxc")
        save_btn.setStyleSheet(
            "QPushButton { background:#ffffff; color:#3a3a5c;"
            "  border:1px solid #d6c0c5; border-radius:5px;"
            "  padding:7px 18px; font-size:12px; font-weight:600; }"
            "QPushButton:hover { background:#f4f6fa; }"
        )
        save_btn.clicked.connect(self.save_data)
        bar_row.addWidget(save_btn)

        gen_btn = QPushButton("Generate Document")
        gen_btn.setStyleSheet(
            "QPushButton { background:#920d2e; color:#ffffff;"
            "  border:none; border-radius:5px;"
            "  padding:7px 22px; font-size:13px; font-weight:700; }"
            "QPushButton:hover { background:#7a0b27; }"
            "QPushButton:pressed { background:#600820; }"
        )
        gen_btn.clicked.connect(self.generate_cost_sheet)
        bar_row.addWidget(gen_btn)
        root.addWidget(bar)

        # Initial calculation
        self._recalculate()

    def _recalculate(self):
        if not hasattr(self, '_summary'):
            return
        self._summary.update(
            self._consulting.subtotal(),
            self._third_party.subtotal(),
            self._materials.subtotal(),
            self._travel.subtotal(),
        )

    # ── Save / Load ────────────────────────────────────────────────────────
    def get_data(self) -> dict:
        return {
            "client":      getattr(self, "_client", None) and self._client.text() or "",
            "project":     self._proj.text(),
            "date":        self._date.text(),
            "consulting":  self._consulting.get_data(),
            "third_party": self._third_party.get_data(),
            "materials":   self._materials.get_data(),
            "travel":      self._travel.get_data(),
            "summary":     self._summary.get_data(),
        }

    def restore_data(self, d: dict):
        self._client.setText(d.get("client", ""))
        self._proj.setText(d.get("project", ""))
        self._date.setText(d.get("date", ""))
        self._consulting.restore_data(d.get("consulting", {}))
        self._third_party.restore_data(d.get("third_party", {}))
        self._materials.restore_data(d.get("materials", {}))
        self._travel.restore_data(d.get("travel", {}))
        self._summary.restore_data(d.get("summary", {}))
        self._recalculate()

    # ── Generate / Import ──────────────────────────────────────────────────
    def generate_cost_sheet(self):
        """Save .mcmxc + Excel breakdown side-by-side."""
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        client  = self._client.text().strip().replace(" ", "_") or "CostSheet"
        import re as _re
        safe = _re.sub(r'[\\/:*?"<>|]', "-", client)
        default = os.path.join(desktop, f"{safe}.mcmxc")
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Cost Sheet", default,
            "MCMX Cost Sheet (*.mcmxc)")
        if not path:
            return
        # ── .mcmxc (JSON) ─────────────────────────────────────────────────
        data = self.get_data()
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"_type": "mcmxc", "_version": 1, **data}, f,
                          indent=2, ensure_ascii=False)
        except Exception as e:
            QMessageBox.critical(self, "Save Failed", str(e)); return
        # ── Excel ──────────────────────────────────────────────────────────
        xlsx_path = os.path.splitext(path)[0] + ".xlsx"
        try:
            _export_excel(data, xlsx_path)
        except Exception as e:
            QMessageBox.warning(self, "Excel Export Failed", str(e))
            return
        QMessageBox.information(
            self, "Cost Sheet Generated",
            f"✔  Files saved:\n{path}\n{xlsx_path}")

    def save_data(self):
        """Save cost estimator state as .mcmxc (JSON only, no Excel)."""
        import re as _re
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        client  = self._client.text().strip().replace(" ", "_") or "CostSheet"
        safe    = _re.sub(r'[:*?"<>|]', "-", client)
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Cost Sheet",
            os.path.join(desktop, f"{safe}.mcmxc"),
            "MCMX Cost Sheet (*.mcmxc)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"_type": "mcmxc", "_version": 1, **self.get_data()},
                          f, indent=2, ensure_ascii=False)
            QMessageBox.information(self, "Saved", f"✔  Saved:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Save Failed", str(e))

    def import_cost_sheet(self):
        """Load a previously saved .mcmxc file."""
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Cost Sheet", desktop,
            "MCMX Cost Sheet (*.mcmxc)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            QMessageBox.critical(self, "Import Failed", str(e)); return
        self.restore_data(data)
        QMessageBox.information(self, "Imported", "✔  Cost sheet restored.")


# ══════════════════════════════════════════════════════════════════════════════
# Excel export
# ══════════════════════════════════════════════════════════════════════════════
def _export_excel(data: dict, path: str):
    """
    Write a formatted Excel workbook from the cost estimator data dict.
    Requires openpyxl (already a transitive dependency via python-docx).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Breakdown"

    # ── Colour palette ────────────────────────────────────────────────────
    RED     = "920D2E"
    RED_LT  = "F5D0DA"
    GRAY    = "F4F6FA"
    BLUE_BG = "E8F0FE"
    BLUE_FG = "1A3A6E"
    WHITE   = "FFFFFF"
    DARK    = "1A0509"
    BDR_C   = "D6C0C5"

    thin = Side(style="thin", color=BDR_C)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color=DARK, size=11, italic=False):
        return Font(name="Aptos", bold=bold, color=color, size=size, italic=italic)

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    row = 1

    # ── Title bar ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, "MCMX Cost Estimator")
    c.font = _font(bold=True, color=WHITE, size=16)
    c.fill = _fill(RED)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 28
    row += 1

    # ── Project info ──────────────────────────────────────────────────────
    for lbl, key in [("Client", "client"), ("Project", "project"), ("Date", "date")]:
        ws.cell(row, 1, lbl).font = _font(bold=True)
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row, 2, data.get(key, "")).font = _font()
        ws.row_dimensions[row].height = 18
        row += 1
    row += 1  # blank

    # ── Helpers ───────────────────────────────────────────────────────────
    def _section_hdr(title):
        nonlocal row
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row, 1, title)
        c.font = _font(bold=True, color=WHITE, size=12)
        c.fill = _fill(RED)
        c.alignment = _align("left")
        ws.row_dimensions[row].height = 22
        row += 1

    def _col_hdrs(*labels):
        nonlocal row
        for ci, lbl in enumerate(labels, 1):
            c = ws.cell(row, ci, lbl)
            c.font = _font(bold=True)
            c.fill = _fill(GRAY)
            c.alignment = _align("center")
            c.border = bdr
        ws.row_dimensions[row].height = 18
        row += 1

    def _data_row(*vals, shade=False):
        nonlocal row
        bg = GRAY if shade else WHITE
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row, ci, val)
            c.font = _font()
            c.fill = _fill(bg)
            c.alignment = _align("right" if ci > 1 else "left")
            c.border = bdr
            if isinstance(val, (int, float)) and ci > 1:
                c.number_format = "#,##0"
        ws.row_dimensions[row].height = 17
        row += 1

    def _subtotal_row(val):
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        lc = ws.cell(row, 1, "Subtotal")
        lc.font = _font(bold=True)
        lc.fill = _fill(RED_LT)
        lc.alignment = _align("right")
        lc.border = bdr
        vc = ws.cell(row, 4, val)
        vc.font = _font(bold=True, color=BLUE_FG)
        vc.fill = _fill(BLUE_BG)
        vc.alignment = _align("right")
        vc.border = bdr
        vc.number_format = "#,##0"
        ws.row_dimensions[row].height = 17
        row += 2  # blank after section

    def _summary_row(label, value, accent=False, pct_str=None):
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        lc = ws.cell(row, 1, label)
        lc.font = _font(bold=accent, color=RED if accent else DARK)
        lc.alignment = _align("right")
        lc.border = bdr
        if pct_str is not None:
            vc = ws.cell(row, 4, pct_str)
            vc.number_format = "@"
        else:
            vc = ws.cell(row, 4, value)
            vc.number_format = "#,##0"
        vc.font = _font(bold=accent, color=RED if accent else BLUE_FG)
        vc.fill = _fill(RED_LT if accent else BLUE_BG)
        vc.alignment = _align("right")
        vc.border = bdr
        ws.row_dimensions[row].height = 18
        row += 1

    # ── IBE Level rates lookup ────────────────────────────────────────────
    _IBE_RATES = {1: 7.5, 2: 5.0, 3: 3.375, 4: 1.875}

    # ─────────────────────────────────────────────────────────────────────
    # 1. IBE Estimator (only if active)
    # ─────────────────────────────────────────────────────────────────────
    ibe = data.get("ibe", {})
    if ibe.get("active", False):
        lvl      = ibe.get("level", 1)
        ibe_rate = _IBE_RATES.get(lvl, 7.5)
        panels   = ibe.get("panels", "0")
        hours    = ibe.get("hours", "8")
        n_techs  = ibe.get("num_techs", 1)
        confirmed = ibe.get("confirmed", {})
        insp_days  = confirmed.get("insp_days", "—")
        hotel_nts  = confirmed.get("hotel_nights", "—")
        tot_meals  = confirmed.get("meals", "—")
        flights    = confirmed.get("flights", "—")
        drive_mi   = confirmed.get("drive_miles", "—")

        tech_names  = ibe.get("tech_names",  [])
        tech_modes  = ibe.get("tech_modes",  [])
        tech_travel = ibe.get("tech_travel", [])
        tech_flight = ibe.get("tech_flight_costs", [])
        tech_miles  = ibe.get("tech_mileages",     [])

        _section_hdr("1.  IBE Estimator")

        # Config summary rows
        def _cfg_row(label, value):
            nonlocal row
            ws.cell(row, 1, label).font = _font(bold=True)
            ws.merge_cells(f"B{row}:D{row}")
            ws.cell(row, 2, value).font = _font()
            ws.row_dimensions[row].height = 17
            row += 1

        _cfg_row("IBE Level",     f"Level {lvl}  —  {ibe_rate:g} panels / hr")
        _cfg_row("Total Panels",  panels)
        _cfg_row("Hours / Day",   hours)
        _cfg_row("Technicians",   str(n_techs))
        _cfg_row("Inspection Days",  str(insp_days))
        _cfg_row("Hotel Nights (total)", str(hotel_nts))
        _cfg_row("Total Meals",   str(tot_meals))
        if confirmed.get("flights", 0):
            _cfg_row("Round-Trip Flights", str(flights))
        if confirmed.get("drive_miles", 0):
            _cfg_row("Drive Miles (total)", f"{drive_mi:,.0f}" if isinstance(drive_mi, float) else str(drive_mi))
        row += 1  # blank

        # Tech table
        _col_hdrs("Technician", "Mode", "Travel Time", "Mileage / Flight Cost")
        for ti in range(n_techs):
            nm   = tech_names[ti]  if ti < len(tech_names)  else f"Tech {ti+1}"
            mode = tech_modes[ti]  if ti < len(tech_modes)  else "Driving"
            tt   = tech_travel[ti] if ti < len(tech_travel) else "0"
            if mode == "Flying":
                extra = f"${_num_s(tech_flight[ti] if ti < len(tech_flight) else '0'):,.0f} / trip"
            else:
                extra = f"{_num_s(tech_miles[ti] if ti < len(tech_miles) else '0'):g} mi one-way"
            _data_row(nm, mode, f"{_num_s(tt):g} hrs", extra, shade=(ti % 2 == 1))
        _subtotal_row(0)   # placeholder — IBE cost flows into sections 2 & 5
        row -= 1  # remove extra blank from _subtotal_row since we add another below
        row += 1  # blank after section

    # ─────────────────────────────────────────────────────────────────────
    # 2. Consulting Effort
    # ─────────────────────────────────────────────────────────────────────
    cons = data.get("consulting", {})
    rate = _num_s(cons.get("rate", "0"))
    hours_list = cons.get("rows", [])
    row_rates  = cons.get("row_rates", [])
    cat_labels = ["On-Site Time (hrs)", "Travel Time (hrs)", "Other (hrs)"]

    _section_hdr("2.  Consulting Effort")
    _col_hdrs("Category", "Hours", "Rate ($/hr)", "Cost")
    cons_sub = 0.0
    for i, lbl in enumerate(cat_labels):
        hrs = _num_s(hours_list[i] if i < len(hours_list) else "0")
        rr  = row_rates[i] if i < len(row_rates) else ""
        effective_rate = _num_s(rr) if rr.strip() else rate
        cost = math.ceil(hrs * effective_rate)
        cons_sub += cost
        _data_row(lbl, hrs, effective_rate, cost, shade=(i % 2 == 1))
    _subtotal_row(math.ceil(cons_sub))

    # ─────────────────────────────────────────────────────────────────────
    # 3. Third Party
    # ─────────────────────────────────────────────────────────────────────
    tp = data.get("third_party", {})
    _section_hdr("3.  Third Party")
    _col_hdrs("Description", "Qty", "Cost Per", "Cost Total")
    tp_total = 0.0
    for i, row_vals in enumerate(tp.get("rows", [])):
        desc, qty, cost = (list(row_vals) + ["", "0", "0"])[:3]
        q = _num_s(qty); c = _num_s(cost)
        v = math.ceil(q * c)
        tp_total += v
        _data_row(desc, q, f"${c:,.2f}" if c else "—", v, shade=(i % 2 == 1))
    _subtotal_row(math.ceil(tp_total))

    # ─────────────────────────────────────────────────────────────────────
    # 4. Hardware / Software / Materials
    # ─────────────────────────────────────────────────────────────────────
    mat = data.get("materials", {})
    _section_hdr("4.  Hardware / Software / Materials")
    _col_hdrs("Description", "Qty", "Cost Per", "Cost Total")
    mat_total = 0.0
    for i, row_vals in enumerate(mat.get("rows", [])):
        desc, qty, cost = (list(row_vals) + ["", "0", "0"])[:3]
        q = _num_s(qty); c = _num_s(cost)
        v = math.ceil(q * c)
        mat_total += v
        _data_row(desc, q, c, v, shade=(i % 2 == 1))
    _subtotal_row(math.ceil(mat_total))

    # ─────────────────────────────────────────────────────────────────────
    # 5. Travel & Expenses
    # ─────────────────────────────────────────────────────────────────────
    trav = data.get("travel", {})
    qty_list  = trav.get("qty",  [])
    rate_list = trav.get("rate", [])
    t_labels  = ["Airfare", "Hotel (nights)", "Food (meals)", "Car (miles)"]
    _section_hdr("5.  Travel & Expenses")
    _col_hdrs("Category", "Qty", "Rate", "Cost")
    trav_total = 0.0
    for i, lbl in enumerate(t_labels):
        q = _num_s(qty_list[i]  if i < len(qty_list)  else "0")
        r = _num_s(rate_list[i] if i < len(rate_list) else "0")
        v = math.ceil(q * r)
        trav_total += v
        _data_row(lbl, q, r, v, shade=(i % 2 == 1))
    _subtotal_row(math.ceil(trav_total))

    # ─────────────────────────────────────────────────────────────────────
    # 5. Summary
    # ─────────────────────────────────────────────────────────────────────
    summ = data.get("summary", {})
    risk_pct   = _num_s(summ.get("risk_pct",   "0"))
    margin_pct = _num_s(summ.get("margin_pct", "0"))
    cost_before = math.ceil(cons_sub) + math.ceil(tp_total) + math.ceil(mat_total) + math.ceil(trav_total)
    risk_cost   = math.ceil(cost_before * risk_pct / 100)
    estimated   = cost_before + risk_cost
    resale      = math.ceil(estimated / (1 - margin_pct / 100)) if 0 < margin_pct < 100 else estimated
    profit      = resale - estimated

    _section_hdr("6.  Cost Summary")
    _summary_row("Cost Before Risk",                    cost_before)
    _summary_row(f"Risk / Insurance ({risk_pct:.1f}%)", risk_cost)
    _summary_row("Estimated Cost",                      estimated)
    _summary_row(f"Margin ({margin_pct:.1f}%)",         None, pct_str=f"{margin_pct:.1f}%")
    _summary_row("Recommended Resale",                  resale,  accent=True)
    _summary_row("Profit",                              profit,  accent=True)

    # ─────────────────────────────────────────────────────────────────────
    # IBE Schedule sheet (only if active and schedule present)
    # ─────────────────────────────────────────────────────────────────────
    ibe          = data.get("ibe", {})
    schedule     = ibe.get("schedule", [])
    hotel_states = ibe.get("hotel_states", [])
    n_techs      = ibe.get("num_techs", 1)
    tech_names   = ibe.get("tech_names", [])
    tech_modes   = ibe.get("tech_modes", [])
    tech_travel  = ibe.get("tech_travel", [])
    tech_miles   = ibe.get("tech_mileages", [])
    tech_flight  = ibe.get("tech_flight_costs", [])
    confirmed    = ibe.get("confirmed", {})

    # Pull rates from the Travel card (index 0=air, 1=hotel, 2=meal, 3=mile)
    trav_rates   = data.get("travel", {}).get("rate", [])
    air_rate     = _num_s(trav_rates[0]) if len(trav_rates) > 0 else 450.0
    hotel_rate   = _num_s(trav_rates[1]) if len(trav_rates) > 1 else 150.0
    meal_rate    = _num_s(trav_rates[2]) if len(trav_rates) > 2 else 25.0
    mile_rate    = _num_s(trav_rates[3]) if len(trav_rates) > 3 else 0.67

    if ibe.get("active", False) and schedule:
        ws2 = wb.create_sheet("IBE Schedule")
        from openpyxl.utils import get_column_letter as _gcl

        def _sf(bold=False, color=DARK, size=11, italic=False):
            return Font(name="Aptos", bold=bold, color=color, size=size, italic=italic)
        def _sfill(hex_c):
            return PatternFill("solid", fgColor=hex_c)
        def _salign(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        # Per-tech alternating header colours (header bg / data bg)
        _TECH_PALETTES = [
            ("920d2e", "FFF0F3"),   # dark red / blush
            ("1F4E79", "EBF3FB"),   # dark blue / sky blue
            ("375623", "EBF5E8"),   # dark green / mint
            ("7B3F00", "FFF3E8"),   # dark orange / peach
            ("4B0082", "F3EBF9"),   # indigo / lavender
        ]
        def _tech_hdr_fill(ti):  return _TECH_PALETTES[ti % len(_TECH_PALETTES)][0]
        def _tech_data_fill(ti): return _TECH_PALETTES[ti % len(_TECH_PALETTES)][1]

        # Column layout: Day(1) | Activity(2) | Panels Done(3) | % Done(4) | [per tech: 7 cols]
        # Per-tech: Travel(hrs) | Miles / Transport | Mileage Cost | Hotel? | Hotel Cost | Meals | Meals Cost
        PER_T      = 7
        BASE       = 5
        TOTAL_COLS = 4 + n_techs * PER_T
        SUB_LABELS = ["Travel (hrs)", "Miles", "Mileage $", "Hotel?", "Hotel Cost", "Meals", "Meals Cost"]
        sub_widths = [13,             14,       11,          8,        11,            7,       11          ]

        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 28
        ws2.column_dimensions["C"].width = 13
        ws2.column_dimensions["D"].width = 11
        for ti in range(n_techs):
            for ci in range(PER_T):
                ws2.column_dimensions[_gcl(BASE + ti * PER_T + ci)].width = sub_widths[ci]

        sr = 1

        # Title
        ws2.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=TOTAL_COLS)
        tc = ws2.cell(sr, 1, "IBE Work Week Schedule")
        tc.font = _sf(bold=True, color=WHITE, size=14)
        tc.fill = _sfill(RED); tc.alignment = _salign("center")
        ws2.row_dimensions[sr].height = 26; sr += 1

        # Config
        ibe_lvl = ibe.get("level", 1)
        ws2.cell(sr, 1, f"IBE Level {ibe_lvl}").font = _sf(bold=True)
        ws2.cell(sr, 2,
            f"{ibe.get('panels','?')} panels   {ibe.get('hours','?')} hrs/day   {n_techs} technician(s)"
        ).font = _sf()
        ws2.row_dimensions[sr].height = 16; sr += 1

        # Rates
        ws2.cell(sr, 1, "Rates used:").font = _sf(bold=True)
        ws2.cell(sr, 2,
            f"Hotel ${hotel_rate:,.2f}/night     Meals ${meal_rate:,.2f}/meal"
            f"     Mileage ${mile_rate:,.3f}/mi     Airfare ${air_rate:,.0f}/trip"
        ).font = _sf(italic=True, color="555555")
        ws2.row_dimensions[sr].height = 16; sr += 2

        # Tech name header row
        for col, lbl in [(1,"Day"),(2,"Activity"),(3,"Panels Done"),(4,"% Done")]:
            c = ws2.cell(sr, col, lbl)
            c.font = _sf(bold=True); c.fill = _sfill(GRAY); c.alignment = _salign("center", wrap=True)
        for ti in range(n_techs):
            nm = tech_names[ti] if ti < len(tech_names) else f"Tech {ti+1}"
            sc = BASE + ti * PER_T
            ws2.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + PER_T - 1)
            nc = ws2.cell(sr, sc, nm)
            nc.font = _sf(bold=True, color=WHITE)
            nc.fill = _sfill(_tech_hdr_fill(ti))
            nc.alignment = _salign("center")
        ws2.row_dimensions[sr].height = 22; sr += 1

        # Sub-header row
        for col in (1,2,3,4):
            ws2.cell(sr, col).fill = _sfill(GRAY)
        for ti in range(n_techs):
            sc = BASE + ti * PER_T
            for ci, lbl in enumerate(SUB_LABELS):
                c = ws2.cell(sr, sc + ci, lbl)
                c.font = _sf(bold=True, size=10)
                c.fill = _sfill(_tech_hdr_fill(ti))
                c.alignment = _salign("center", wrap=True)
                c.font = Font(name="Aptos", bold=True, color=WHITE, size=10)
        ws2.row_dimensions[sr].height = 30; sr += 1

        # Data helpers
        total_panels_f      = _num_s(ibe.get("panels", "0"))
        ibe_rate_v          = {1: 7.5, 2: 5.0, 3: 3.375, 4: 1.875}.get(ibe_lvl, 7.5)
        hours_day_f         = _num_s(ibe.get("hours", "8"))
        panels_per_insp_day = ibe_rate_v * hours_day_f * n_techs
        cum_panels          = 0.0

        def _blank_acc():
            return [{"t_hrs": 0.0, "miles": 0.0, "mile_cost": 0.0,
                     "hotel_nights": 0, "hotel_cost": 0.0,
                     "meals": 0, "meals_cost": 0.0} for _ in range(n_techs)]

        week_acc  = _blank_acc()
        grand_acc = _blank_acc()
        current_week  = 0
        prev_was_weekend = False

        def _write_week_banner(num):
            nonlocal sr
            ws2.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=TOTAL_COLS)
            wc = ws2.cell(sr, 1, f"WEEK {num}")
            wc.font = _sf(bold=True, color=WHITE); wc.fill = _sfill("5B6FA8")
            wc.alignment = _salign("center")
            ws2.row_dimensions[sr].height = 16; sr += 1

        def _write_subtotal(label, acc, fill_hex, txt_color=DARK):
            nonlocal sr
            ws2.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=4)
            lc = ws2.cell(sr, 1, label)
            lc.font = _sf(bold=True, color=txt_color)
            lc.fill = _sfill(fill_hex); lc.alignment = _salign("center")
            for col in (2, 3, 4):
                ws2.cell(sr, col).fill = _sfill(fill_hex)
            for ti in range(n_techs):
                a  = acc[ti]
                sc = BASE + ti * PER_T
                vals = [
                    f"{a['t_hrs']:.1f} hrs"      if a['t_hrs']        else "—",
                    f"{a['miles']:,.0f} mi"       if a['miles']        else "—",
                    f"${a['mile_cost']:,.2f}"     if a['mile_cost']    else "—",
                    str(a['hotel_nights'])         if a['hotel_nights'] else "—",
                    f"${a['hotel_cost']:,.2f}"    if a['hotel_cost']   else "—",
                    str(a['meals'])               if a['meals']        else "—",
                    f"${a['meals_cost']:,.2f}"    if a['meals_cost']   else "—",
                ]
                for ci, v in enumerate(vals):
                    c = ws2.cell(sr, sc + ci, v)
                    c.font = _sf(bold=True, color=txt_color)
                    c.fill = _sfill(fill_hex); c.alignment = _salign("center")
            ws2.row_dimensions[sr].height = 18; sr += 1

        for ri, row_data in enumerate(schedule):
            day_lbl, activity, is_insp, def_hotel = (list(row_data) + ["","",False,False])[:4]
            is_weekend    = (str(day_lbl) == "Weekend")
            is_travel_in  = (activity == "Travel In")
            is_travel_out = (activity == "Travel Out")
            is_travel     = is_travel_in or is_travel_out

            if is_weekend:
                _write_subtotal(f"Week {current_week} Subtotal", week_acc, "D9E1F2", "1F4E79")
                week_acc = _blank_acc()

            if ri == 0:
                current_week = 1
                _write_week_banner(current_week)
            elif not is_weekend and prev_was_weekend:
                current_week += 1
                _write_week_banner(current_week)

            prev_was_weekend = is_weekend
            row_bg = "F2F4F8" if is_weekend else WHITE

            # Day cell
            dc = ws2.cell(sr, 1, str(day_lbl))
            dc.font = _sf(bold=not is_weekend, color=("666666" if is_weekend else DARK))
            dc.fill = _sfill(row_bg); dc.alignment = _salign("center")

            # Activity cell
            ac = ws2.cell(sr, 2, str(activity))
            ac.font = _sf(italic=is_weekend, color=("888888" if is_weekend else DARK))
            ac.fill = _sfill(row_bg)

            # Panels Done + % Done
            if is_insp:
                cum_panels = min(cum_panels + panels_per_insp_day, total_panels_f)
                pct = (cum_panels / total_panels_f * 100) if total_panels_f > 0 else 0
                pd_cell  = ws2.cell(sr, 3, f"{int(round(cum_panels)):,}")
                pct_cell = ws2.cell(sr, 4, f"{min(pct, 100.0):.1f}%")
                pd_cell.font  = _sf(bold=True, color=RED)
                pct_cell.font = _sf(bold=True, color=RED)
            else:
                pd_cell  = ws2.cell(sr, 3, "—")
                pct_cell = ws2.cell(sr, 4, "—")
                pd_cell.font  = _sf(color="AAAAAA")
                pct_cell.font = _sf(color="AAAAAA")
            pd_cell.fill  = _sfill(row_bg); pd_cell.alignment  = _salign("center")
            pct_cell.fill = _sfill(row_bg); pct_cell.alignment = _salign("center")

            # Per-tech columns
            for ti in range(n_techs):
                sc     = BASE + ti * PER_T
                mode   = tech_modes[ti]  if ti < len(tech_modes)  else "Driving"
                ow_mi  = _num_s(tech_miles[ti]  if ti < len(tech_miles)  else "0")
                f_cost = _num_s(tech_flight[ti] if ti < len(tech_flight) else "0") or air_rate
                t_hrs  = _num_s(tech_travel[ti] if ti < len(tech_travel) else "0")
                flying = (mode == "Flying")
                data_bg = _tech_data_fill(ti)

                hotel_booked = (
                    bool(hotel_states[ti][ri])
                    if ti < len(hotel_states) and ri < len(hotel_states[ti])
                    else bool(def_hotel)
                )

                if is_weekend:
                    rt_hrs    = t_hrs * 2
                    rt_mi     = ow_mi * 2 if not flying else 0.0
                    mi_cost   = rt_mi * mile_rate if not flying else f_cost * 2
                    mi_str    = (f"{rt_mi:,.0f} mi RT" if rt_mi else "—") if not flying else "Flying RT"
                    cost_str  = f"${mi_cost:,.2f}" if mi_cost else "—"

                    vals = [
                        f"{rt_hrs:.1f} hrs" if rt_hrs else "—",
                        mi_str, cost_str,
                        "Home", "—", "—", "—",
                    ]
                    week_acc[ti]["t_hrs"]     += rt_hrs
                    week_acc[ti]["miles"]     += rt_mi
                    week_acc[ti]["mile_cost"] += mi_cost
                    grand_acc[ti]["t_hrs"]    += rt_hrs
                    grand_acc[ti]["miles"]    += rt_mi
                    grand_acc[ti]["mile_cost"]+= mi_cost
                    fills = ["F2F4F8"] * PER_T

                else:
                    if is_travel:
                        row_t   = t_hrs
                        row_mi  = ow_mi if not flying else 0.0
                    elif is_insp and not hotel_booked and not flying:
                        row_t   = t_hrs * 2
                        row_mi  = ow_mi * 2
                    else:
                        row_t   = 0.0
                        row_mi  = 0.0

                    mi_cost   = row_mi * mile_rate
                    n_meals   = 2
                    h_cost    = hotel_rate if hotel_booked else 0.0
                    m_cost    = n_meals * meal_rate

                    if flying and is_travel:
                        mi_str  = f"Flying  (${f_cost:,.0f}/trip)"
                        mi_cost = f_cost if is_travel_in else 0.0   # charge on inbound only
                    elif row_mi > 0:
                        sfx    = " RT" if (is_insp and not hotel_booked) else " OW"
                        mi_str = f"{row_mi:,.0f} mi{sfx}"
                    else:
                        mi_str = "—"

                    hotel_str = "Yes" if hotel_booked else "No"
                    h_str     = f"${h_cost:,.2f}"    if h_cost    else "—"
                    mc_str    = f"${mi_cost:,.2f}"   if mi_cost   else "—"

                    vals = [
                        f"{row_t:.1f} hrs" if row_t else "—",
                        mi_str, mc_str,
                        hotel_str, h_str,
                        str(n_meals), f"${m_cost:,.2f}",
                    ]
                    fills = [
                        data_bg, data_bg, data_bg,
                        ("FFF0F3" if hotel_booked else data_bg),
                        ("FFF0F3" if hotel_booked else data_bg),
                        data_bg, data_bg,
                    ]

                    week_acc[ti]["t_hrs"]       += row_t
                    week_acc[ti]["miles"]       += row_mi
                    week_acc[ti]["mile_cost"]   += mi_cost
                    week_acc[ti]["hotel_nights"]+= (1 if hotel_booked else 0)
                    week_acc[ti]["hotel_cost"]  += h_cost
                    week_acc[ti]["meals"]       += n_meals
                    week_acc[ti]["meals_cost"]  += m_cost
                    grand_acc[ti]["t_hrs"]       += row_t
                    grand_acc[ti]["miles"]        += row_mi
                    grand_acc[ti]["mile_cost"]    += mi_cost
                    grand_acc[ti]["hotel_nights"] += (1 if hotel_booked else 0)
                    grand_acc[ti]["hotel_cost"]   += h_cost
                    grand_acc[ti]["meals"]        += n_meals
                    grand_acc[ti]["meals_cost"]   += m_cost

                for ci, (v, sf) in enumerate(zip(vals, fills)):
                    c = ws2.cell(sr, sc + ci, v)
                    c.font      = _sf(color="666666" if is_weekend else DARK)
                    c.fill      = _sfill(sf)
                    c.alignment = _salign("center")

            ws2.row_dimensions[sr].height = 18; sr += 1

        _write_subtotal(f"Week {current_week} Subtotal", week_acc, "D9E1F2", "1F4E79")

        sr += 1
        _write_subtotal("GRAND TOTAL", grand_acc, RED, WHITE)

        # Confirmed totals summary
        sr += 2
        ws2.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=TOTAL_COLS)
        hdr2 = ws2.cell(sr, 1, "Confirmed Schedule Totals")
        hdr2.font = _sf(bold=True, color=WHITE); hdr2.fill = _sfill(RED)
        hdr2.alignment = _salign("center")
        ws2.row_dimensions[sr].height = 20; sr += 1

        tot_rows = [
            ("Inspection Days",  str(confirmed.get("insp_days",    "—"))),
            ("Hotel Nights",     str(confirmed.get("hotel_nights", "—"))),
            ("Total Meals",      str(confirmed.get("meals",        "—"))),
            ("Drive Miles",      f"{confirmed.get('drive_miles', 0):,.0f}" if confirmed.get("drive_miles") else "—"),
            ("Flights",          str(confirmed.get("flights",  "—"))),
            ("On-Site Hours",    str(confirmed.get("onsite_hrs","—"))),
            ("Total Travel Hrs", str(confirmed.get("travel_hrs","—"))),
        ]
        for lbl, val in tot_rows:
            ws2.cell(sr, 1, lbl).font = _sf(bold=True)
            ws2.cell(sr, 2, val).font = _sf()
            ws2.row_dimensions[sr].height = 16; sr += 1

    wb.save(path)


def _num_s(s) -> float:
    """Parse a numeric string from saved cost data."""
    try:
        return float(str(s).replace(",", "").replace("$", "").strip() or 0)
    except (ValueError, TypeError):
        return 0.0
