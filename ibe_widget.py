# ibe_widget.py
"""
IBE (Inspection-Based Estimator) widget.

Computes on-site days from IBE level, panel count, hours/day, and technician count.
Generates a work-week overview with per-tech hotel checkboxes, then pre-populates
the Cost Estimator's Consulting and Travel cards when confirmed.

IBEWidget  — standalone tab widget (wraps content in a scroll area)
IBECard    — collapsible _Card for embedding directly in the Cost Estimator
"""

import math
import json
import os
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QScrollArea, QFrame, QSizePolicy, QComboBox,
    QCheckBox, QAbstractItemView, QFileDialog, QMessageBox,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QDoubleValidator, QIntValidator

# ── Palette (mirrors cost_estimator.py) ─────────────────────────────────────
_RED     = "#920d2e"
_TEXT    = "#1a0509"
_SUBTEXT = "#3a3a5c"
_BG      = "#f0f2f7"
_BG_CARD = "#ffffff"
_BG_HDR  = "#f9f0f2"
_BORDER  = "#d6c0c5"
_BLUE_BG = "#e8f0fe"
_BLUE_FG = "#1a3a6e"

_FIELD_STYLE = (
    "QLineEdit { color:#1a0509; background:#ffffff;"
    "  border:1px solid #d6c0c5; border-radius:4px;"
    "  padding:4px 8px; font-size:12px; }"
    "QLineEdit:focus { border-color:#920d2e; }"
)
_CALC_STYLE = (
    "QLineEdit { color:#1a3a6e; background:#e8f0fe;"
    "  border:1px solid #b8cce4; border-radius:4px;"
    "  padding:4px 8px; font-size:12px; font-weight:600; }"
)
_BTN_STYLE = (
    "QPushButton { background:#f4f6fa; color:#3a3a5c;"
    "  border:1px solid #dde1e7; border-radius:4px;"
    "  padding:4px 10px; font-size:12px; }"
    "QPushButton:hover { background:#e8ecf5; border-color:#b0b8d0; }"
    "QPushButton:pressed { background:#d8def0; }"
)
_ACTN_STYLE = (
    "QPushButton { background:#920d2e; color:#ffffff; border:none;"
    "  border-radius:5px; padding:7px 22px; font-size:13px; font-weight:700; }"
    "QPushButton:hover { background:#7a0b27; }"
    "QPushButton:pressed { background:#600820; }"
    "QPushButton:disabled { background:#c8c8c8; color:#888888; }"
)
_SPIN_S = (
    "QPushButton { background:#f4f6fa; color:#3a3a5c; border:1px solid #dde1e7;"
    "  border-radius:4px; font-size:14px; font-weight:700; padding:0; }"
    "QPushButton:hover { background:#e8ecf5; }"
)
_CB_STYLE = (
    "QCheckBox { spacing:6px; font-size:11px; color:#1a0509; }"
    "QCheckBox::indicator { width:16px; height:16px; border-radius:3px; }"
    "QCheckBox::indicator:unchecked { border:2px solid #d6c0c5; background:#ffffff; }"
    "QCheckBox::indicator:unchecked:hover { border-color:#920d2e; }"
    "QCheckBox::indicator:checked { border:2px solid #920d2e; background:#920d2e; }"
)
_INFO_LBL = f"font-size:10px; color:{_SUBTEXT}; background:transparent; border:none;"
_OUTL_STYLE = (
    "QPushButton { background:#ffffff; color:#920d2e;"
    "  border:1px solid #d6c0c5; border-radius:5px;"
    "  padding:7px 18px; font-size:12px; font-weight:600; }"
    "QPushButton:hover { background:#fdf0f3; border-color:#920d2e; }"
)
_TABLE_STYLE = (
    "QTableWidget { background:#ffffff; border:1px solid #d6c0c5;"
    "  border-radius:4px; gridline-color:#e8dde0; font-size:11px; }"
    "QHeaderView::section { background:#f4f6fa; color:#3a3a5c;"
    "  font-size:11px; font-weight:600; border:none;"
    "  border-bottom:1px solid #d6c0c5; padding:4px 6px; }"
    "QTableWidget::item { padding:3px 6px; }"
    "QTableWidget::item:selected { background:#f5d0da; color:#920d2e; }"
)

# Panel inspection rates by IBE level
_LEVEL_RATES = {1: 7.5, 2: 5.0, 3: 3.375, 4: 1.875}
_DAY_NAMES   = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _n(text, default=0.0):
    try:
        return float(str(text).replace(",", "").replace("$", "").strip() or default)
    except (ValueError, TypeError):
        return default


# ── Shared collapsible card (same visual as cost_estimator) ─────────────────
class _Card(QFrame):
    def __init__(self, title, parent=None):
        super().__init__(parent)
        self.setStyleSheet(
            f"QFrame {{ background:{_BG_CARD}; border:1px solid {_BORDER}; border-radius:8px; }}"
        )
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        hdr = QWidget()
        hdr.setStyleSheet(
            f"QWidget {{ background:{_BG_HDR}; border-radius:7px 7px 0 0;"
            f"  border-bottom:1px solid {_BORDER}; }}"
        )
        hdr.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        hdr.setFixedHeight(40)
        self._hdr_row = QHBoxLayout(hdr)
        self._hdr_row.setContentsMargins(12, 0, 8, 0)
        self._hdr_row.setSpacing(8)

        self._toggle_btn = QPushButton("▼")
        self._toggle_btn.setFixedSize(22, 22)
        self._toggle_btn.setStyleSheet(
            f"QPushButton {{ background:transparent; border:none;"
            f"  color:{_RED}; font-size:12px; font-weight:700; padding:0; }}"
        )
        self._toggle_btn.clicked.connect(self._toggle)
        self._hdr_row.addWidget(self._toggle_btn)

        self._title_lbl = QLabel(title)
        self._title_lbl.setStyleSheet(
            f"QLabel {{ font-size:13px; font-weight:700; color:{_RED};"
            "  background:transparent; border:none; padding:0; }"
        )
        self._hdr_row.addWidget(self._title_lbl)
        self._hdr_row.addStretch()

        self._badge = QLabel("")
        self._badge.setStyleSheet(
            f"QLabel {{ font-size:12px; font-weight:700; color:{_RED};"
            "  background:transparent; border:none; padding:0 4px; }"
        )
        self._hdr_row.addWidget(self._badge)
        root.addWidget(hdr)

        self._body = QWidget()
        self._body.setStyleSheet("QWidget { background:transparent; border:none; }")
        self._body_layout = QVBoxLayout(self._body)
        self._body_layout.setContentsMargins(14, 12, 14, 14)
        self._body_layout.setSpacing(8)
        root.addWidget(self._body)

        self._collapsed = False

    def _toggle(self):
        self._collapsed = not self._collapsed
        self._body.setVisible(not self._collapsed)
        self._toggle_btn.setText("►" if self._collapsed else "▼")

    def set_badge(self, text):
        self._badge.setText(text)


# ── Main IBE widget ──────────────────────────────────────────────────────────
class IBEWidget(QWidget):
    """
    Drop-in IBE widget.
    compact=False  → standalone tab with scroll area
    compact=True   → no scroll wrapper (for embedding inside IBECard)
    """

    LABOR_RATE  = 125.0
    TRAVEL_RATE = 100.0
    HOTEL_RATE  = 150.0
    MEAL_RATE   =  25.0
    MILE_RATE   =   0.72

    def __init__(self, parent=None, compact=False):
        super().__init__(parent)
        self._schedule        = []
        self._hotel_checks    = []   # [tech_idx][row_idx] → QCheckBox | None
        self._tech_rows       = []
        self._num_techs       = 1
        self._confirmed       = {}
        self._row_hours_edits = {}   # row_idx → QLineEdit for per-day hour override

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        if compact:
            cl = QVBoxLayout()
            cl.setContentsMargins(0, 0, 0, 0)
            cl.setSpacing(8)
            root.addLayout(cl)
        else:
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setFrameShape(QFrame.Shape.NoFrame)
            scroll.setStyleSheet(
                f"QScrollArea {{ background:{_BG}; border:none; }}"
                "QScrollBar:vertical { background:#f0f2f7; width:8px; margin:0; }"
                "QScrollBar::handle:vertical { background:#c8cedd; border-radius:4px; min-height:20px; }"
                "QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0; }"
            )
            content = QWidget()
            content.setStyleSheet(f"QWidget {{ background:{_BG}; }}")
            cl = QVBoxLayout(content)
            cl.setContentsMargins(20, 16, 20, 16)
            cl.setSpacing(12)

        # Configuration card
        self._cfg_card = _Card("IBE Configuration")
        self._build_config(self._cfg_card._body_layout)
        cl.addWidget(self._cfg_card)

        # Work week overview card (hidden until generated)
        self._outlook_card = _Card("Work Week Overview")
        self._build_outlook_card(self._outlook_card._body_layout)
        self._outlook_card.setVisible(False)
        cl.addWidget(self._outlook_card)

        # Schedule summary card (hidden until confirmed)
        self._summary_card = _Card("Schedule Summary")
        self._build_summary_card(self._summary_card._body_layout)
        self._summary_card.setVisible(False)
        cl.addWidget(self._summary_card)

        cl.addStretch()

        if not compact:
            scroll.setWidget(content)
            root.addWidget(scroll)

            # ── Bottom action bar ────────────────────────────────────────
            _bar = QWidget()
            _bar.setStyleSheet("QWidget { background:#f9f0f2; border-top:1px solid #d6c0c5; }")
            _bar.setFixedHeight(52)
            _br = QHBoxLayout(_bar)
            _br.setContentsMargins(16, 8, 16, 8)
            _br.setSpacing(10)

            _br.addStretch()

            _load_btn = QPushButton("⬆  Import .mcmxi…")
            _load_btn.setStyleSheet(
                "QPushButton { background:#ffffff; color:#920d2e;"
                "  border:1px solid #d6c0c5; border-radius:5px;"
                "  padding:7px 18px; font-size:12px; font-weight:600; }"
                "QPushButton:hover { background:#fdf0f3; border-color:#920d2e; }"
            )
            _load_btn.clicked.connect(self.load_data)
            _br.addWidget(_load_btn)

            _save_btn = QPushButton("💾  Save .mcmxi")
            _save_btn.setStyleSheet(
                "QPushButton { background:#ffffff; color:#3a3a5c;"
                "  border:1px solid #d6c0c5; border-radius:5px;"
                "  padding:7px 18px; font-size:12px; font-weight:600; }"
                "QPushButton:hover { background:#f4f6fa; }"
            )
            _save_btn.clicked.connect(self.save_data)
            _br.addWidget(_save_btn)

            self._export_btn = QPushButton("📊  Export Excel…")
            self._export_btn.setStyleSheet(
                "QPushButton { background:#920d2e; color:#ffffff;"
                "  border:none; border-radius:5px;"
                "  padding:7px 18px; font-size:13px; font-weight:700; }"
                "QPushButton:hover { background:#7a0b27; }"
                "QPushButton:pressed { background:#600820; }"
                "QPushButton:disabled { background:#c8c8c8; color:#888888; }"
            )
            self._export_btn.setEnabled(False)
            self._export_btn.setToolTip("Confirm the schedule first.")
            self._export_btn.clicked.connect(self.export_excel)
            _br.addWidget(self._export_btn)

            root.addWidget(_bar)

    # ── File I/O ──────────────────────────────────────────────────────────────

    def save_data(self):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        path, _ = QFileDialog.getSaveFileName(
            self, "Save IBE Project",
            os.path.join(desktop, "ibe.mcmxi"),
            "IBE Project (*.mcmxi)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.get_data(), f, indent=2, ensure_ascii=False)
            QMessageBox.information(self, "Saved", f"✔  Saved:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Save Failed", str(e))

    def load_data(self):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        path, _ = QFileDialog.getOpenFileName(
            self, "Load IBE Project", desktop,
            "IBE Project (*.mcmxi)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                self.restore_data(json.load(f))
        except Exception as e:
            QMessageBox.critical(self, "Load Failed", str(e))

    def export_excel(self):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        path, _ = QFileDialog.getSaveFileName(
            self, "Export IBE Schedule to Excel",
            os.path.join(desktop, "ibe_schedule.xlsx"),
            "Excel Workbook (*.xlsx)")
        if not path:
            return
        try:
            _data = self.get_data()
            _export_ibe_excel(_data, path,
                              labor_rate=self.LABOR_RATE,
                              travel_rate=self.TRAVEL_RATE,
                              hotel_rate=self.HOTEL_RATE,
                              meal_rate=self.MEAL_RATE,
                              mile_rate=self.MILE_RATE)
            _mcmxi_path = os.path.splitext(path)[0] + ".mcmxi"
            try:
                with open(_mcmxi_path, "w", encoding="utf-8") as _mf:
                    json.dump(_data, _mf, indent=2, ensure_ascii=False)
                QMessageBox.information(
                    self, "Exported",
                    f"✔  Excel saved:\n{path}\n\nProject saved:\n{_mcmxi_path}")
            except Exception as _me:
                QMessageBox.warning(
                    self, "Project Save Warning",
                    f"Excel exported but .mcmxi could not be saved:\n{_me}")
                QMessageBox.information(self, "Exported", f"✔  Excel saved:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))

    # ── Config card ──────────────────────────────────────────────────────────
    def _build_config(self, layout):
        # Level selection
        lev_row = QHBoxLayout()
        lev_row.setSpacing(12)
        lev_lbl = QLabel("IBE Level:")
        lev_lbl.setStyleSheet(f"font-size:12px; font-weight:700; color:{_RED};")
        lev_lbl.setFixedWidth(78)
        lev_row.addWidget(lev_lbl)

        self._level_combo = QComboBox()
        self._level_combo.setStyleSheet(
            "QComboBox {"
            "  font-size:12px; color:#1a0509; background:#ffffff;"
            "  border:1px solid #d6c0c5; border-radius:4px;"
            "  padding:4px 10px; min-width:260px; }"
            "QComboBox:focus { border-color:#920d2e; }"
            "QComboBox:hover { border-color:#920d2e; }"
            "QComboBox::drop-down {"
            "  subcontrol-origin:padding; subcontrol-position:top right;"
            "  width:28px; border-left:1px solid #d6c0c5;"
            "  border-top-right-radius:4px; border-bottom-right-radius:4px;"
            "  background:#f4f6fa; }"
            "QComboBox::down-arrow {"
            "  image:none; width:0; height:0;"
            "  border-left:5px solid transparent;"
            "  border-right:5px solid transparent;"
            "  border-top:6px solid #920d2e; }"
            "QComboBox QAbstractItemView {"
            "  font-size:12px; color:#1a0509; background:#ffffff;"
            "  border:1px solid #d6c0c5; border-radius:4px;"
            "  selection-background-color:#f5d0da;"
            "  selection-color:#920d2e;"
            "  outline:none; }"
        )
        for lvl, rate in _LEVEL_RATES.items():
            self._level_combo.addItem(f"Level {lvl}  —  {rate:g} panels / hr", lvl)
        self._level_combo.currentIndexChanged.connect(self._recalc_preview)
        lev_row.addWidget(self._level_combo)
        lev_row.addStretch()
        layout.addLayout(lev_row)

        # Panels + hours per day
        ph_row = QHBoxLayout()
        ph_row.setSpacing(12)

        pl = QLabel("Total Panels:")
        pl.setStyleSheet(f"font-size:12px; font-weight:700; color:{_RED};")
        ph_row.addWidget(pl)
        self._panels_edit = QLineEdit()
        self._panels_edit.setPlaceholderText("e.g. 500")
        self._panels_edit.setValidator(QIntValidator(0, 9_999_999))
        self._panels_edit.setFixedWidth(110)
        self._panels_edit.setStyleSheet(_FIELD_STYLE)
        self._panels_edit.textChanged.connect(self._recalc_preview)
        self._panels_edit.textChanged.connect(self._update_gen_btn_state)
        ph_row.addWidget(self._panels_edit)

        ph_row.addSpacing(16)
        hl = QLabel("Hours / Day:")
        hl.setStyleSheet(f"font-size:12px; font-weight:700; color:{_RED};")
        ph_row.addWidget(hl)
        self._hours_edit = QLineEdit("8")
        self._hours_edit.setValidator(QDoubleValidator(0.5, 24.0, 1))
        self._hours_edit.setFixedWidth(70)
        self._hours_edit.setStyleSheet(_FIELD_STYLE)
        self._hours_edit.textChanged.connect(self._recalc_preview)
        ph_row.addWidget(self._hours_edit)
        ph_row.addStretch()
        layout.addLayout(ph_row)

        # Technician count
        tc_row = QHBoxLayout()
        tc_row.setSpacing(8)
        tl = QLabel("Technicians:")
        tl.setStyleSheet(f"font-size:12px; font-weight:700; color:{_RED};")
        tl.setFixedWidth(90)
        tc_row.addWidget(tl)

        self._tech_minus = QPushButton("−")
        self._tech_minus.setFixedSize(28, 28)
        self._tech_minus.setStyleSheet(_SPIN_S)
        self._tech_minus.clicked.connect(self._dec_techs)
        tc_row.addWidget(self._tech_minus)

        self._tech_count_lbl = QLabel("1")
        self._tech_count_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._tech_count_lbl.setStyleSheet(
            f"font-size:13px; font-weight:700; color:{_TEXT}; min-width:24px;"
        )
        tc_row.addWidget(self._tech_count_lbl)

        self._tech_plus = QPushButton("+")
        self._tech_plus.setFixedSize(28, 28)
        self._tech_plus.setStyleSheet(_SPIN_S)
        self._tech_plus.clicked.connect(self._inc_techs)
        tc_row.addWidget(self._tech_plus)
        tc_row.addStretch()
        layout.addLayout(tc_row)

        # Technician entry rows
        self._tech_container = QWidget()
        self._tech_container.setStyleSheet("background:transparent;")
        self._tech_vbox = QVBoxLayout(self._tech_container)
        self._tech_vbox.setContentsMargins(0, 4, 0, 0)
        self._tech_vbox.setSpacing(6)
        layout.addWidget(self._tech_container)
        self._rebuild_tech_rows()

        # Day-of-week checkboxes
        dw_row = QHBoxLayout()
        dw_row.setSpacing(8)
        dw_lbl = QLabel("Work Days:")
        dw_lbl.setStyleSheet(f"font-size:12px; font-weight:700; color:{_RED};")
        dw_lbl.setFixedWidth(78)
        dw_row.addWidget(dw_lbl)
        self._day_checks = []
        defaults = [True, True, True, True, True, False, False]
        for name, default in zip(_DAY_NAMES, defaults):
            cb = QCheckBox(name)
            cb.setChecked(default)
            cb.setStyleSheet(f"QCheckBox {{ font-size:12px; color:{_TEXT}; }}")
            cb.stateChanged.connect(self._recalc_preview)
            cb.toggled.connect(self._mark_outlook_stale)
            dw_row.addWidget(cb)
            self._day_checks.append(cb)
        dw_row.addStretch()
        layout.addLayout(dw_row)

        # Start Day of Week + Work Hours
        sw_row = QHBoxLayout()
        sw_row.setSpacing(16)

        sd_lbl = QLabel("Start Day:")
        sd_lbl.setStyleSheet(f"font-size:12px; font-weight:700; color:{_RED};")
        sd_lbl.setFixedWidth(78)
        sw_row.addWidget(sd_lbl)

        self._start_day_combo = QComboBox()
        self._start_day_combo.setStyleSheet(
            "QComboBox {"
            "  font-size:12px; color:#1a0509; background:#ffffff;"
            "  border:1px solid #d6c0c5; border-radius:4px;"
            "  padding:4px 10px; min-width:100px; }"
            "QComboBox:focus { border-color:#920d2e; }"
            "QComboBox:hover { border-color:#920d2e; }"
            "QComboBox::drop-down {"
            "  subcontrol-origin:padding; subcontrol-position:top right;"
            "  width:28px; border-left:1px solid #d6c0c5;"
            "  border-top-right-radius:4px; border-bottom-right-radius:4px;"
            "  background:#f4f6fa; }"
            "QComboBox::down-arrow {"
            "  image:none; width:0; height:0;"
            "  border-left:5px solid transparent;"
            "  border-right:5px solid transparent;"
            "  border-top:6px solid #920d2e; }"
            "QComboBox QAbstractItemView {"
            "  font-size:12px; color:#1a0509; background:#ffffff;"
            "  border:1px solid #d6c0c5; border-radius:4px;"
            "  selection-background-color:#f5d0da;"
            "  selection-color:#920d2e;"
            "  outline:none; }"
        )
        for _dn in _DAY_NAMES:
            self._start_day_combo.addItem(_dn)
        self._start_day_combo.setCurrentIndex(0)
        self._start_day_combo.setToolTip("First day of the first inspection week")
        self._start_day_combo.currentIndexChanged.connect(self._mark_outlook_stale)
        sw_row.addWidget(self._start_day_combo)

        sw_row.addSpacing(24)

        wh_lbl = QLabel("Work Hours:")
        wh_lbl.setStyleSheet(f"font-size:12px; font-weight:700; color:{_RED};")
        sw_row.addWidget(wh_lbl)

        _TIME_COMBO_STYLE = (
            "QComboBox {"
            "  font-size:11px; color:#1a0509; background:#ffffff;"
            "  border:1px solid #d6c0c5; border-radius:4px;"
            "  padding:3px 8px; min-width:90px; }"
            "QComboBox:focus { border-color:#920d2e; }"
            "QComboBox:hover { border-color:#920d2e; }"
            "QComboBox::drop-down {"
            "  subcontrol-origin:padding; subcontrol-position:top right;"
            "  width:22px; border-left:1px solid #d6c0c5;"
            "  border-top-right-radius:4px; border-bottom-right-radius:4px;"
            "  background:#f4f6fa; }"
            "QComboBox::down-arrow {"
            "  image:none; width:0; height:0;"
            "  border-left:5px solid transparent;"
            "  border-right:5px solid transparent;"
            "  border-top:6px solid #920d2e; }"
            "QComboBox QAbstractItemView {"
            "  font-size:11px; color:#1a0509; background:#ffffff;"
            "  border:1px solid #d6c0c5; selection-background-color:#f5d0da;"
            "  selection-color:#920d2e; outline:none; }"
        )

        def _hour_lbl(h):
            if h == 0:  return "12:00 AM"
            if h < 12:  return f"{h}:00 AM"
            if h == 12: return "12:00 PM"
            return f"{h - 12}:00 PM"

        self._work_start_combo = QComboBox()
        self._work_start_combo.setStyleSheet(_TIME_COMBO_STYLE)
        for _h in range(24):
            self._work_start_combo.addItem(_hour_lbl(_h), _h)
        self._work_start_combo.setCurrentIndex(8)   # 8:00 AM
        self._work_start_combo.setToolTip(
            "Shift start time.\n"
            "Weekday hours before this are overtime (× 1.5).")
        sw_row.addWidget(self._work_start_combo)

        wh_to = QLabel("to")
        wh_to.setStyleSheet(f"font-size:12px; color:{_TEXT};")
        sw_row.addWidget(wh_to)

        self._work_end_combo = QComboBox()
        self._work_end_combo.setStyleSheet(_TIME_COMBO_STYLE)
        for _h in range(24):
            self._work_end_combo.addItem(_hour_lbl(_h), _h)
        self._work_end_combo.setCurrentIndex(17)    # 5:00 PM
        self._work_end_combo.setToolTip(
            "Shift end time.\n"
            "Weekday hours after this are overtime (× 1.5).\n"
            "Set end before start for overnight shifts (e.g. 10:00 PM → 8:00 AM).")
        sw_row.addWidget(self._work_end_combo)

        self._wh_window_lbl = QLabel("")
        self._wh_window_lbl.setStyleSheet(f"font-size:10px; color:{_SUBTEXT}; font-style:italic;")
        sw_row.addWidget(self._wh_window_lbl)
        self._work_start_combo.currentIndexChanged.connect(self._update_wh_window_lbl)
        self._work_end_combo.currentIndexChanged.connect(self._update_wh_window_lbl)
        self._update_wh_window_lbl()

        sw_row.addStretch()
        layout.addLayout(sw_row)

        # Preview strip
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"color:{_BORDER};")
        layout.addWidget(sep)

        prev_row = QHBoxLayout()
        prev_row.setSpacing(32)
        for attr, caption in [
            ("_prev_pd",    "Panels / Day"),
            ("_prev_days",  "Inspection Days"),
            ("_prev_weeks", "Approx. Weeks"),
        ]:
            col = QVBoxLayout()
            col.setSpacing(2)
            lbl = QLabel(caption)
            lbl.setStyleSheet(f"font-size:10px; color:{_SUBTEXT};")
            col.addWidget(lbl)
            val = QLabel("—")
            val.setStyleSheet(f"font-size:20px; font-weight:800; color:{_RED};")
            col.addWidget(val)
            setattr(self, attr, val)
            prev_row.addLayout(col)
        prev_row.addStretch()
        layout.addLayout(prev_row)

        # Generate Work Week Overview button
        gen_row = QHBoxLayout()
        gen_row.addStretch()
        self._gen_btn = QPushButton("Generate Work Week Overview  →")
        self._gen_btn.setStyleSheet(_ACTN_STYLE)
        self._gen_btn.setEnabled(False)
        self._gen_btn.setToolTip("Enter the total panel count first.")
        self._gen_btn.clicked.connect(self._generate_outlook)
        gen_row.addWidget(self._gen_btn)
        layout.addLayout(gen_row)

    # ── Outlook card ─────────────────────────────────────────────────────────
    def _build_outlook_card(self, layout):
        self._outlook_table = QTableWidget()
        self._outlook_table.setStyleSheet(_TABLE_STYLE)
        self._outlook_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._outlook_table.verticalHeader().setVisible(False)
        self._outlook_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self._outlook_table)

        confirm_row = QHBoxLayout()
        confirm_row.addStretch()
        self._confirm_btn = QPushButton("Confirm Schedule  →")
        self._confirm_btn.setStyleSheet(_ACTN_STYLE)
        self._confirm_btn.setEnabled(False)
        self._confirm_btn.setToolTip("Generate the Work Week Overview first.")
        self._confirm_btn.clicked.connect(self._confirm_schedule)
        confirm_row.addWidget(self._confirm_btn)
        layout.addLayout(confirm_row)

    # ── Summary card ─────────────────────────────────────────────────────────
    def _build_summary_card(self, layout):
        # Stats row
        stats_row = QHBoxLayout(); stats_row.setSpacing(20)
        for attr, caption in [
            ("_sum_insp",    "Inspection Days"),
            ("_sum_hotel",   "Hotel Nights"),
            ("_sum_meals",   "Total Meals"),
            ("_sum_flights", "Air Trips"),
            ("_sum_miles",   "Drive Miles"),
        ]:
            col = QVBoxLayout(); col.setSpacing(2)
            lbl = QLabel(caption); lbl.setStyleSheet(f"font-size:10px; color:{_SUBTEXT};")
            col.addWidget(lbl)
            val = QLineEdit("—"); val.setReadOnly(True)
            val.setStyleSheet(_CALC_STYLE); val.setFixedWidth(100)
            col.addWidget(val)
            setattr(self, attr, val)
            stats_row.addLayout(col)
        stats_row.addStretch()
        layout.addLayout(stats_row)

        # Meals count is manually adjustable
        self._sum_meals.setReadOnly(False)
        self._sum_meals.setStyleSheet(
            "QLineEdit { color:#1a3a6e; background:#ffffff;"
            "  border:2px solid #7b8cde; border-radius:4px;"
            "  padding:4px 8px; font-size:14px; font-weight:700; }"
            "QLineEdit:focus { border-color:#920d2e; }"
        )
        self._sum_meals.setToolTip("Total Meals — edit to manually override")
        self._sum_meals.editingFinished.connect(self._on_meals_edited)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"color:{_BORDER}; margin-top:4px;"); layout.addWidget(sep)

        # Cost breakdown
        cost_lbl = QLabel("Cost Breakdown")
        cost_lbl.setStyleSheet(f"font-size:12px; font-weight:700; color:{_RED}; margin-top:4px;")
        layout.addWidget(cost_lbl)

        _FORMULA_STYLE = f"font-size:9px; color:{_SUBTEXT}; font-style:italic;"
        cost_grid = QHBoxLayout(); cost_grid.setSpacing(20)
        for attr, caption in [
            ("_cost_labor",   "Labor (Work)"),
            ("_cost_travel",  "Labor (Travel)"),
            ("_cost_hotel",   "Hotel"),
            ("_cost_meals",   "Meals"),
            ("_cost_miles",   "Mileage"),
            ("_cost_flights", "Flights"),
        ]:
            col = QVBoxLayout(); col.setSpacing(2)
            lbl = QLabel(caption); lbl.setStyleSheet(f"font-size:10px; color:{_SUBTEXT};")
            col.addWidget(lbl)
            val = QLineEdit("—"); val.setReadOnly(True)
            val.setStyleSheet(_CALC_STYLE); val.setFixedWidth(100)
            col.addWidget(val)
            setattr(self, attr, val)
            formula = QLabel("")
            formula.setStyleSheet(_FORMULA_STYLE)
            formula.setWordWrap(True)
            formula.setMaximumWidth(130)
            col.addWidget(formula)
            setattr(self, attr + "_formula", formula)
            cost_grid.addLayout(col)
        cost_grid.addStretch()
        layout.addLayout(cost_grid)

        # Margin control + totals
        margin_row = QHBoxLayout(); margin_row.setSpacing(12)

        ml = QLabel("Margin %:")
        ml.setStyleSheet(f"font-size:12px; font-weight:700; color:{_RED};")
        margin_row.addWidget(ml)

        self._margin_minus = QPushButton("−")
        self._margin_minus.setFixedSize(28, 28)
        self._margin_minus.setStyleSheet(_SPIN_S)
        self._margin_minus.clicked.connect(lambda: self._adj_margin(-1))
        margin_row.addWidget(self._margin_minus)

        self._margin_lbl = QLineEdit("0.0")
        self._margin_lbl.setFixedWidth(60)
        self._margin_lbl.setStyleSheet(_FIELD_STYLE)
        self._margin_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._margin_lbl.setValidator(QDoubleValidator(0.0, 99.9, 1))
        self._margin_lbl.textChanged.connect(self._recalc_cost_totals)
        margin_row.addWidget(self._margin_lbl)

        pct_lbl = QLabel("%")
        pct_lbl.setStyleSheet(f"font-size:12px; color:{_TEXT};")
        margin_row.addWidget(pct_lbl)

        self._margin_plus = QPushButton("+")
        self._margin_plus.setFixedSize(28, 28)
        self._margin_plus.setStyleSheet(_SPIN_S)
        self._margin_plus.clicked.connect(lambda: self._adj_margin(+1))
        margin_row.addWidget(self._margin_plus)

        margin_row.addSpacing(24)

        for attr, caption, is_grand in [
            ("_cost_subtotal", "Subtotal (before margin)", False),
            ("_cost_grand",    "Grand Total",              True),
        ]:
            lbl = QLabel(caption + ":")
            lbl.setStyleSheet(f"font-size:12px; color:{_SUBTEXT};")
            margin_row.addWidget(lbl)
            val = QLineEdit("—"); val.setReadOnly(True)
            val.setStyleSheet(
                _CALC_STYLE if not is_grand else
                "QLineEdit { color:#920d2e; background:#f9f0f2;"
                "  border:1px solid #d6c0c5; border-radius:4px;"
                "  padding:4px 8px; font-size:13px; font-weight:700; }"
            )
            val.setFixedWidth(120 if is_grand else 110)
            setattr(self, attr, val)
            margin_row.addWidget(val)
        margin_row.addStretch()
        layout.addLayout(margin_row)

    # ── Technician row management ─────────────────────────────────────────────
    def _rebuild_tech_rows(self):
        while self._tech_vbox.count():
            item = self._tech_vbox.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._tech_rows = []

        _mode_style = (
            "QComboBox { font-size:11px; color:#1a0509; background:#ffffff;"
            "  border:1px solid #d6c0c5; border-radius:4px; padding:3px 8px; }"
            "QComboBox:focus { border-color:#920d2e; }"
            "QComboBox::drop-down { border:none; width:18px; background:#f4f6fa;"
            "  border-left:1px solid #d6c0c5; border-top-right-radius:4px;"
            "  border-bottom-right-radius:4px; }"
            "QComboBox::down-arrow { width:0; height:0;"
            "  border-left:4px solid transparent; border-right:4px solid transparent;"
            "  border-top:5px solid #920d2e; }"
            "QComboBox QAbstractItemView { font-size:11px; color:#1a0509;"
            "  background:#ffffff; border:1px solid #d6c0c5;"
            "  selection-background-color:#f5d0da; selection-color:#920d2e; }"
        )

        for i in range(self._num_techs):
            row_w = QWidget()
            row_w.setStyleSheet("background:transparent;")
            hl = QHBoxLayout(row_w)
            hl.setContentsMargins(0, 0, 0, 0)
            hl.setSpacing(8)

            num_lbl = QLabel(f"Tech {i + 1}:")
            num_lbl.setStyleSheet(f"font-size:12px; color:{_SUBTEXT}; min-width:52px;")
            hl.addWidget(num_lbl)

            name_ed = QLineEdit()
            name_ed.setPlaceholderText(f"Technician {i + 1}")
            name_ed.setStyleSheet(_FIELD_STYLE)
            name_ed.setFixedWidth(150)
            hl.addWidget(name_ed)

            tl = QLabel("Travel Time:")
            tl.setStyleSheet(f"font-size:12px; color:{_SUBTEXT};")
            hl.addWidget(tl)

            travel_ed = QLineEdit()
            travel_ed.setPlaceholderText("hrs")
            travel_ed.setValidator(QDoubleValidator(0.0, 48.0, 1))
            travel_ed.setFixedWidth(60)
            travel_ed.setStyleSheet(_FIELD_STYLE)
            travel_ed.textChanged.connect(self._mark_outlook_stale)
            hl.addWidget(travel_ed)

            # Transport mode — Driving default
            mode_combo = QComboBox()
            mode_combo.addItem("Driving", "driving")
            mode_combo.addItem("Flying",  "flying")
            mode_combo.setFixedWidth(90)
            mode_combo.setStyleSheet(_mode_style)
            mode_combo.currentTextChanged.connect(self._mark_outlook_stale)
            hl.addWidget(mode_combo)

            # Flight cost field (shown when Flying)
            cost_lbl = QLabel("Flight Cost:")
            cost_lbl.setStyleSheet(f"font-size:12px; color:{_SUBTEXT};")
            hl.addWidget(cost_lbl)
            cost_ed = QLineEdit()
            cost_ed.setPlaceholderText("$ per trip")
            cost_ed.setValidator(QDoubleValidator(0.0, 99999.0, 2))
            cost_ed.setFixedWidth(90)
            cost_ed.setStyleSheet(_FIELD_STYLE)
            hl.addWidget(cost_ed)

            # Mileage field (shown when Driving)
            mile_lbl = QLabel("Mileage:")
            mile_lbl.setStyleSheet(f"font-size:12px; color:{_SUBTEXT};")
            hl.addWidget(mile_lbl)
            mile_ed = QLineEdit()
            mile_ed.setPlaceholderText("mi one-way")
            mile_ed.setValidator(QDoubleValidator(0.0, 99999.0, 1))
            mile_ed.setFixedWidth(90)
            mile_ed.setStyleSheet(_FIELD_STYLE)
            hl.addWidget(mile_ed)

            hl.addStretch()

            def _toggle(text, cl=cost_lbl, ce=cost_ed, ml=mile_lbl, me=mile_ed):
                flying = text == "Flying"
                cl.setVisible(flying)
                ce.setVisible(flying)
                ml.setVisible(not flying)
                me.setVisible(not flying)

            mode_combo.currentTextChanged.connect(_toggle)
            _toggle(mode_combo.currentText())   # apply initial state (Driving)

            self._tech_vbox.addWidget(row_w)
            self._tech_rows.append({
                "name":        name_ed,
                "travel":      travel_ed,
                "mode":        mode_combo,
                "flight_cost": cost_ed,
                "mileage":     mile_ed,
            })

    def _snapshot_tech_data(self) -> list:
        """Capture current tech row values before a rebuild."""
        saved = []
        for row in self._tech_rows:
            saved.append({
                "name":        row["name"].text(),
                "travel":      row["travel"].text(),
                "mode":        row["mode"].currentText(),
                "flight_cost": row["flight_cost"].text(),
                "mileage":     row["mileage"].text(),
            })
        return saved

    def _restore_tech_data(self, saved: list):
        """Restore tech row values after a rebuild (preserves existing techs)."""
        for i, s in enumerate(saved):
            if i >= len(self._tech_rows):
                break
            row = self._tech_rows[i]
            row["name"].setText(s["name"])
            row["travel"].setText(s["travel"])
            idx = row["mode"].findText(s["mode"])
            if idx >= 0:
                row["mode"].setCurrentIndex(idx)
            row["flight_cost"].setText(s["flight_cost"])
            row["mileage"].setText(s["mileage"])

    def _inc_techs(self):
        if self._num_techs < 10:
            saved = self._snapshot_tech_data()
            self._num_techs += 1
            self._tech_count_lbl.setText(str(self._num_techs))
            self._rebuild_tech_rows()
            self._restore_tech_data(saved)
            self._recalc_preview()
            self._mark_outlook_stale()

    def _dec_techs(self):
        if self._num_techs > 1:
            saved = self._snapshot_tech_data()
            self._num_techs -= 1
            self._tech_count_lbl.setText(str(self._num_techs))
            self._rebuild_tech_rows()
            self._restore_tech_data(saved)
            self._recalc_preview()
            self._mark_outlook_stale()

    # ── Preview ───────────────────────────────────────────────────────────────
    def _get_rate(self):
        lvl = self._level_combo.currentData()
        return _LEVEL_RATES.get(lvl, 7.5)

    def _get_work_days(self):
        return [i for i, cb in enumerate(self._day_checks) if cb.isChecked()]

    def _recalc_preview(self):
        rate   = self._get_rate()
        panels = _n(self._panels_edit.text())
        hours  = _n(self._hours_edit.text(), 8.0)
        n      = self._num_techs
        wdays  = self._get_work_days()
        dpw    = len(wdays) or 1

        daily = rate * hours * n
        if daily > 0 and panels > 0:
            days  = math.ceil(panels / daily)
            weeks = math.ceil(days / dpw)
            self._prev_pd.setText(f"{daily:,.2f}")
            self._prev_days.setText(str(days))
            self._prev_weeks.setText(str(weeks))
        else:
            self._prev_pd.setText("—")
            self._prev_days.setText("—")
            self._prev_weeks.setText("—")

    # ── Generate overview ─────────────────────────────────────────────────────
    def _generate_outlook(self):
        rate   = self._get_rate()
        panels = _n(self._panels_edit.text())
        hours  = _n(self._hours_edit.text(), 8.0)
        n      = self._num_techs
        wdays  = self._get_work_days()

        if panels <= 0 or hours <= 0 or not wdays:
            return

        self._row_hours_edits = {}  # reset per-day hour overrides

        daily     = rate * hours * n
        insp_days = math.ceil(panels / daily) if daily > 0 else 0

        # Gap days between end of one work week and start of the next
        # e.g. Mon-Fri → 2 gap days (Sat+Sun); Mon-Sun → 0 gap days
        if len(wdays) > 1:
            gap_days = (wdays[0] + 7 - wdays[-1] - 1) % 7
        else:
            gap_days = 6  # only one work day per week → 6 gap days
        self._weekend_days = gap_days

        # First week may start mid-week based on the chosen Start Day
        start_day_idx = (self._start_day_combo.currentIndex()
                         if hasattr(self, "_start_day_combo") else 0)
        first_week_days = [wd for wd in wdays if wd >= start_day_idx]
        if not first_week_days:
            first_week_days = wdays  # fall back if start is after all work days

        # Build schedule: (day_label, activity, is_inspection, default_hotel)
        sched = []

        # Travel In/Out only when at least one tech's journey warrants it
        # (flying, or drive time >= 2 hrs one-way).  Local techs drive RT
        # each day — no overnight stays, so no dedicated travel days.
        _any_hotel_needed = any(
            _n(self._tech_rows[i]["travel"].text(), 0.0) >= 2.0 or
            self._tech_rows[i]["mode"].currentText() == "Flying"
            for i in range(n)
            if i < len(self._tech_rows)
        )

        # Travel In: day before first inspection day
        travel_in_day = _DAY_NAMES[(first_week_days[0] - 1) % 7]
        if _any_hotel_needed:
            sched.append((travel_in_day, "Travel In", False, True))

        done       = 0
        last_wd    = first_week_days[0]
        first_week = True
        while done < insp_days:
            week_days = first_week_days if first_week else wdays
            if not first_week and gap_days > 0:
                sched.append(("Weekend", "— Travel Home / Return —", False, False))
            first_week = False
            for wd in week_days:
                if done >= insp_days:
                    break
                done    += 1
                last_wd  = wd
                sched.append((_DAY_NAMES[wd], f"Inspection  Day {done}", True, True))

        # Travel Out: day after last work day
        travel_out_day = _DAY_NAMES[(last_wd + 1) % 7]
        if _any_hotel_needed:
            sched.append((travel_out_day, "Travel Out", False, False))

        self._schedule  = sched
        self._confirmed = {}
        self._refresh_outlook_table(sched, n)
        self._outlook_card.setVisible(True)
        self._summary_card.setVisible(False)
        self._confirm_btn.setEnabled(True)
        self._confirm_btn.setToolTip("")

    def _refresh_outlook_table(self, sched, num_techs):
        tech_names = [
            self._tech_rows[i]["name"].text().strip() or f"Tech {i + 1}"
            for i in range(num_techs)
        ]
        tech_info = []
        for i in range(num_techs):
            if i < len(self._tech_rows):
                row = self._tech_rows[i]
                tech_info.append({
                    "mode":        row["mode"].currentText(),
                    "travel_hrs":  _n(row["travel"].text(), 0.0),
                    "mileage_ow":  _n(row["mileage"].text(), 0.0),
                    "flight_cost": _n(row["flight_cost"].text(), 0.0),
                })
            else:
                tech_info.append({"mode": "Driving", "travel_hrs": 0.0,
                                   "mileage_ow": 0.0, "flight_cost": 0.0})

        headers = ["Day", "Activity", "Hrs"] + tech_names
        self._outlook_table.setColumnCount(len(headers))
        self._outlook_table.setHorizontalHeaderLabels(headers)
        self._outlook_table.setRowCount(len(sched))

        hh = self._outlook_table.horizontalHeader()
        hh.setMinimumHeight(46)
        hh.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter | Qt.TextFlag.TextWordWrap)
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self._outlook_table.setColumnWidth(0, 95)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self._outlook_table.setColumnWidth(2, 64)
        for c in range(3, len(headers)):
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.Fixed)
            self._outlook_table.setColumnWidth(c, 155)

        self._hotel_checks = [[None] * len(sched) for _ in range(num_techs)]

        # Default hours for each inspection row = work window duration,
        # falling back to the Hours/Day field if the window is 0 or undefined.
        if hasattr(self, "_work_start_combo") and hasattr(self, "_work_end_combo"):
            _ws = self._work_start_combo.currentIndex()
            _we = self._work_end_combo.currentIndex()
            if _we > _ws:   _win_hrs = _we - _ws
            elif _we < _ws: _win_hrs = 24 - _ws + _we
            else:           _win_hrs = 0
            _default_hrs_text = str(_win_hrs) if _win_hrs > 0 else (self._hours_edit.text() or "8")
        else:
            _default_hrs_text = self._hours_edit.text() or "8"

        for r, (day_lbl, activity, is_insp, def_hotel) in enumerate(sched):
            is_weekend    = day_lbl == "Weekend"
            is_travel_in  = activity == "Travel In"
            is_travel_out = activity == "Travel Out"
            is_travel     = is_travel_in or is_travel_out

            # ── Hrs column (col 2) ────────────────────────────────────────────
            hrs_w = QWidget()
            hrs_w.setStyleSheet("background:#f4f6fa;" if is_weekend else "background:#ffffff;")
            hrs_l = QVBoxLayout(hrs_w)
            hrs_l.setContentsMargins(4, 4, 4, 4)
            hrs_l.setAlignment(Qt.AlignmentFlag.AlignCenter)
            if is_insp:
                he = QLineEdit(_default_hrs_text)
                he.setFixedWidth(46)
                he.setAlignment(Qt.AlignmentFlag.AlignCenter)
                he.setValidator(QDoubleValidator(0.5, 24.0, 1))
                he.setStyleSheet(_FIELD_STYLE)
                he.setToolTip("Hours worked this day (overrides default)")
                hrs_l.addWidget(he)
                self._row_hours_edits[r] = he
            else:
                _dash = QLabel("—")
                _dash.setAlignment(Qt.AlignmentFlag.AlignCenter)
                _dash.setStyleSheet(f"font-size:12px; color:{_SUBTEXT};")
                hrs_l.addWidget(_dash)
            self._outlook_table.setCellWidget(r, 2, hrs_w)

            day_item = QTableWidgetItem(day_lbl)
            day_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            day_item.setForeground(QColor(_SUBTEXT if is_weekend else (_RED if is_insp else _TEXT)))
            if is_weekend:
                day_item.setBackground(QColor("#f4f6fa"))
            self._outlook_table.setItem(r, 0, day_item)

            act_item = QTableWidgetItem(activity)
            act_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            act_item.setForeground(QColor(_SUBTEXT if is_weekend else _TEXT))
            if is_weekend:
                act_item.setBackground(QColor("#f4f6fa"))
                act_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._outlook_table.setItem(r, 1, act_item)

            for t in range(num_techs):
                info   = tech_info[t]
                flying = info["mode"] == "Flying"
                t_hrs  = info["travel_hrs"]
                mi_ow  = info["mileage_ow"]
                effective_def_hotel = def_hotel and (flying or t_hrs >= 2.0)

                cell_w  = QWidget()
                cell_vl = QVBoxLayout(cell_w)
                cell_vl.setContentsMargins(8, 4, 8, 4)
                cell_vl.setSpacing(2)
                cell_vl.setAlignment(Qt.AlignmentFlag.AlignVCenter)

                if is_weekend:
                    cell_w.setStyleSheet("background:#f4f6fa;")
                    if flying:
                        home_txt = "Flying home  (RT)"
                    else:
                        rt = mi_ow * 2
                        home_txt = f"Drive home  ({rt:,.0f} mi RT)" if rt > 0 else "Drive home  (RT)"
                    lbl_h = QLabel(home_txt)
                    lbl_h.setStyleSheet(_INFO_LBL)
                    lbl_h.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    cell_vl.addWidget(lbl_h)
                    if t_hrs > 0:
                        lbl_t = QLabel(f"{t_hrs * 2:.1f} hrs travel")
                        lbl_t.setStyleSheet(_INFO_LBL)
                        lbl_t.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        cell_vl.addWidget(lbl_t)

                elif is_travel:
                    # Hotel checkbox always shown on travel days (no "Travels in day before" toggle)
                    cell_w.setStyleSheet("background:#fff0f3;" if effective_def_hotel else "background:#ffffff;")

                    this_hotel_cb = QCheckBox("Book hotel")
                    this_hotel_cb.setChecked(effective_def_hotel)
                    this_hotel_cb.setStyleSheet(_CB_STYLE)
                    this_hotel_cb.toggled.connect(
                        lambda checked, w=cell_w: w.setStyleSheet(
                            "background:#fff0f3;" if checked else "background:#ffffff;"
                        )
                    )
                    cell_vl.addWidget(this_hotel_cb)
                    self._hotel_checks[t][r] = this_hotel_cb

                    if t_hrs > 0:
                        _tl = QLabel(f"Travel: {t_hrs:.1f} hrs OW")
                        _tl.setStyleSheet(_INFO_LBL)
                        cell_vl.addWidget(_tl)

                    if flying:
                        f_cost = info["flight_cost"]
                        _ml = QLabel(f"Flight: {'${:,.0f}'.format(f_cost) if f_cost > 0 else '—'}")
                    else:
                        if is_travel_out:
                            _ml = QLabel(f"Drive home: {mi_ow:,.0f} mi OW" if mi_ow > 0 else "Drive home: OW")
                        else:
                            _ml = QLabel(f"Drive: {mi_ow:,.0f} mi OW" if mi_ow > 0 else "Drive: — mi")
                    _ml.setStyleSheet(_INFO_LBL)
                    cell_vl.addWidget(_ml)

                else:
                    # Inspection day
                    _has_hotel = effective_def_hotel
                    cell_w.setStyleSheet("background:#fff0f3;" if _has_hotel else "background:#ffffff;")

                    cb = QCheckBox("Book hotel")
                    cb.setChecked(_has_hotel)
                    cb.setStyleSheet(_CB_STYLE)
                    cell_vl.addWidget(cb)
                    self._hotel_checks[t][r] = cb

                    rate_lbl = QLabel(
                        f"${self.LABOR_RATE:.0f}/hr  ×  {_n(self._hours_edit.text(), 8.0):.0f}h"
                    )
                    rate_lbl.setStyleSheet(_INFO_LBL)
                    cell_vl.addWidget(rate_lbl)

                    meals_lbl = QLabel("2 meals")
                    meals_lbl.setStyleSheet(_INFO_LBL)
                    cell_vl.addWidget(meals_lbl)

                    _ow_mi = mi_ow
                    _ow_hr = t_hrs
                    _rt_mi = mi_ow * 2
                    _rt_hr = t_hrs * 2

                    # Non-hotel day: OW home if previous row had hotel, RT otherwise
                    # Look back past Weekend rows (they have no hotel checkbox)
                    _prev_cb = None
                    for _prev_r in range(r - 1, -1, -1):
                        if self._hotel_checks[t][_prev_r] is not None:
                            _prev_cb = self._hotel_checks[t][_prev_r]
                            break
                    _from_hotel = _prev_cb is not None and _prev_cb.isChecked()
                    if _from_hotel:
                        _drive_lbl = QLabel(f"Drive home: {_ow_mi:,.0f} mi OW")
                        _time_lbl  = QLabel(f"Travel home: {_ow_hr:.1f} hrs OW")
                    else:
                        _drive_lbl = QLabel(f"Drive: {_rt_mi:,.0f} mi RT")
                        _time_lbl  = QLabel(f"Travel: {_rt_hr:.1f} hrs RT")
                    _drive_lbl.setStyleSheet(_INFO_LBL)
                    _drive_lbl.setVisible(not _has_hotel and not flying and (_rt_mi > 0 or _ow_mi > 0))
                    cell_vl.addWidget(_drive_lbl)

                    _time_lbl.setStyleSheet(_INFO_LBL)
                    _time_lbl.setVisible(not _has_hotel and (_rt_hr > 0 or _ow_hr > 0))
                    cell_vl.addWidget(_time_lbl)

                    def _on_hotel_toggled(checked, _w=cell_w, _dl=_drive_lbl, _tl=_time_lbl,
                                          _fly=flying, _ow_m=_ow_mi, _rt_m=_rt_mi,
                                          _ow_h=_ow_hr, _rt_h=_rt_hr, _pc=_prev_cb):
                        _w.setStyleSheet("background:#fff0f3;" if checked else "background:#ffffff;")
                        _from_h = _pc is not None and _pc.isChecked()
                        if checked:
                            _dl.setVisible(False)
                            _tl.setVisible(False)
                        else:
                            if _from_h:
                                _dl.setText(f"Drive home: {_ow_m:,.0f} mi OW")
                                _tl.setText(f"Travel home: {_ow_h:.1f} hrs OW")
                            else:
                                _dl.setText(f"Drive: {_rt_m:,.0f} mi RT")
                                _tl.setText(f"Travel: {_rt_h:.1f} hrs RT")
                            _dl.setVisible(not _fly and (_rt_m > 0 or _ow_m > 0))
                            _tl.setVisible(_rt_h > 0 or _ow_h > 0)
                    cb.toggled.connect(_on_hotel_toggled)

                    # When the PREVIOUS row's hotel changes, update this row's labels
                    if _prev_cb is not None:
                        def _on_prev_toggled(prev_checked, _dl=_drive_lbl, _tl=_time_lbl,
                                             _fly=flying, _ow_m=_ow_mi, _rt_m=_rt_mi,
                                             _ow_h=_ow_hr, _rt_h=_rt_hr, _cur_cb=cb):
                            if not _cur_cb.isChecked():
                                if prev_checked:
                                    _dl.setText(f"Drive home: {_ow_m:,.0f} mi OW")
                                    _tl.setText(f"Travel home: {_ow_h:.1f} hrs OW")
                                    _dl.setVisible(not _fly and _ow_m > 0)
                                    _tl.setVisible(_ow_h > 0)
                                else:
                                    _dl.setText(f"Drive: {_rt_m:,.0f} mi RT")
                                    _tl.setText(f"Travel: {_rt_h:.1f} hrs RT")
                                    _dl.setVisible(not _fly and _rt_m > 0)
                                    _tl.setVisible(_rt_h > 0)
                        _prev_cb.toggled.connect(_on_prev_toggled)

                cell_vl.addStretch()
                self._outlook_table.setCellWidget(r, 3 + t, cell_w)

        ROW_H = 110
        for r in range(len(sched)):
            self._outlook_table.setRowHeight(r, ROW_H)
        self._outlook_table.setMinimumHeight(ROW_H * len(sched) + 50)
        self._outlook_table.setMaximumHeight(16_777_215)

    # ── Confirm schedule ──────────────────────────────────────────────────────
    def _confirm_schedule(self):
        if not self._schedule:
            return
        if len(self._hotel_checks) < self._num_techs:
            QMessageBox.warning(self, "Stale Schedule",
                "The technician count changed since the overview was generated.\n"
                "Please regenerate the Work Week Overview first.")
            return

        n            = self._num_techs
        insp_days    = sum(1 for _, _, is_insp, _ in self._schedule if is_insp)
        weekend_rows = [r for r, (dl, _, _, _) in enumerate(self._schedule) if dl == "Weekend"]

        # Hotel nights
        total_hotel = 0
        for t in range(n):
            for cb in self._hotel_checks[t]:
                if cb is not None and cb.isChecked():
                    total_hotel += 1

        # Meals: 2 per on-site row per tech (travel rows + inspection rows, no weekends)
        on_site_rows = sum(1 for dl, _, _, _ in self._schedule if dl != "Weekend")
        total_meals  = on_site_rows * 2 * n

        # Special row indices for hotel-based travel logic
        _ti_idx    = next((r for r, (_, act, _, _) in enumerate(self._schedule) if act == "Travel In"),  None)
        _insp_idxs = [r for r, (_, _, is_insp, _) in enumerate(self._schedule) if is_insp]
        _last_insp = _insp_idxs[-1] if _insp_idxs else None

        # Miles / flights — driven by hotel checkboxes
        flights = 0; total_drive_miles = 0.0
        for i in range(min(n, len(self._tech_rows))):
            row  = self._tech_rows[i]
            mode = row["mode"].currentText()
            hc   = self._hotel_checks[i] if i < len(self._hotel_checks) else []

            def _hv(r_idx, _hc=hc):
                cb = _hc[r_idx] if 0 <= r_idx < len(_hc) else None
                return cb is not None and cb.isChecked()

            ti_hotel   = _hv(_ti_idx)    if _ti_idx   is not None else False
            last_hotel = _hv(_last_insp) if _last_insp is not None else False

            if mode == "Flying":
                if ti_hotel:   flights += 1
                if last_hotel: flights += 1
            else:
                mi_ow = _n(row["mileage"].text())
                dm    = 0.0
                if ti_hotel:
                    dm += mi_ow
                for r_idx, (dl, _, is_insp, _) in enumerate(self._schedule):
                    if is_insp:
                        prev = _hv(r_idx - 1) if r_idx > 0 else False
                        here = _hv(r_idx)
                        if not prev and not here:  dm += mi_ow * 2
                        elif prev and not here:    dm += mi_ow
                        elif not prev and here:    dm += mi_ow
                    elif dl == "Weekend":
                        dm += mi_ow * 2
                if last_hotel:
                    dm += mi_ow
                total_drive_miles += dm

        # Hours — per-row overrides with weekday OT / weekend OT calculation
        _DAY_IDX      = {name: i for i, name in enumerate(_DAY_NAMES)}
        work_start_h  = (self._work_start_combo.currentIndex()
                         if hasattr(self, "_work_start_combo") else 8)
        work_end_h    = (self._work_end_combo.currentIndex()
                         if hasattr(self, "_work_end_combo")   else 17)
        # Overnight shift: end ≤ start wraps past midnight (e.g. 10 PM → 8 AM = 10 h)
        if work_end_h > work_start_h:
            normal_window = float(work_end_h - work_start_h)
        elif work_end_h < work_start_h:
            normal_window = float(24 - work_start_h + work_end_h)
        else:
            normal_window = 0.0   # same time — every hour is OT
        default_hrs   = _n(self._hours_edit.text(), 8.0)

        regular_hrs = 0.0   # weekday within normal window
        ot_hrs      = 0.0   # weekday outside normal window  (× 1.5)
        weekend_hrs = 0.0   # Saturday / Sunday              (× 2.0)

        for r_idx, (dl, act, is_insp, _) in enumerate(self._schedule):
            if not is_insp:
                continue
            _he = self._row_hours_edits.get(r_idx)
            if _he is not None:
                try:
                    day_hrs = float(_he.text() or str(default_hrs))
                except Exception:
                    day_hrs = default_hrs
            else:
                day_hrs = default_hrs
            is_wknd = _DAY_IDX.get(dl, 0) >= 5   # Sat = 5, Sun = 6
            for _t in range(n):
                if is_wknd:
                    weekend_hrs += day_hrs
                else:
                    regular_hrs += min(day_hrs, normal_window)
                    ot_hrs      += max(0.0, day_hrs - normal_window)

        onsite_hrs = regular_hrs + ot_hrs + weekend_hrs
        travel_hrs    = 0.0
        for i in range(min(n, len(self._tech_rows))):
            t_hrs = _n(self._tech_rows[i]["travel"].text())
            hc    = self._hotel_checks[i] if i < len(self._hotel_checks) else []

            def _hv(r_idx, _hc=hc):
                cb = _hc[r_idx] if 0 <= r_idx < len(_hc) else None
                return cb is not None and cb.isChecked()

            ti_hotel   = _hv(_ti_idx)    if _ti_idx   is not None else False
            last_hotel = _hv(_last_insp) if _last_insp is not None else False

            th = 0.0
            if ti_hotel:
                th += t_hrs
            for r_idx, (dl, _, is_insp, _) in enumerate(self._schedule):
                if is_insp:
                    prev = _hv(r_idx - 1) if r_idx > 0 else False
                    here = _hv(r_idx)
                    if not prev and not here:  th += t_hrs * 2
                    elif prev and not here:    th += t_hrs
                    elif not prev and here:    th += t_hrs
                elif dl == "Weekend":
                    th += t_hrs * 2
            if last_hotel:
                th += t_hrs
            travel_hrs += th

        # Flight costs
        flight_total = 0.0
        for i in range(min(n, len(self._tech_rows))):
            if self._tech_rows[i]["mode"].currentText() == "Flying":
                hc = self._hotel_checks[i] if i < len(self._hotel_checks) else []
                ti_hotel   = (hc[_ti_idx].isChecked()
                              if _ti_idx is not None and _ti_idx < len(hc) and hc[_ti_idx]
                              else False)
                last_hotel = (hc[_last_insp].isChecked()
                              if _last_insp is not None and _last_insp < len(hc) and hc[_last_insp]
                              else False)
                fc = _n(self._tech_rows[i]["flight_cost"].text())
                if ti_hotel:   flight_total += fc
                if last_hotel: flight_total += fc

        # Cost breakdown (OT-aware labor)
        _lr = self.LABOR_RATE
        _tr = self.TRAVEL_RATE
        _hr = self.HOTEL_RATE
        _mr = self.MEAL_RATE
        _mi = self.MILE_RATE
        labor_cost  = (regular_hrs * _lr +
                       ot_hrs      * _lr * 1.5 +
                       weekend_hrs * _lr * 2.0)
        travel_cost = travel_hrs        * _tr
        hotel_cost  = total_hotel       * _hr
        meal_cost   = total_meals       * _mr
        mile_cost   = total_drive_miles * _mi

        self._confirmed = {
            "insp_days":    insp_days,   "hotel_nights":  total_hotel,
            "meals":        total_meals, "flights":        flights,
            "drive_miles":  total_drive_miles, "onsite_hrs": onsite_hrs,
            "travel_hrs":   travel_hrs,
            "regular_hrs":  regular_hrs, "ot_hrs":        ot_hrs,
            "weekend_hrs":  weekend_hrs,
            "labor_cost":   labor_cost,  "travel_cost":   travel_cost,
            "hotel_cost":   hotel_cost,  "meal_cost":     meal_cost,
            "mile_cost":    mile_cost,   "flight_total":  flight_total,
        }

        self._sum_insp.setText(str(insp_days))
        self._sum_hotel.setText(str(total_hotel))
        self._sum_meals.setText(str(total_meals))
        self._sum_flights.setText(str(flights))
        self._sum_miles.setText(f"{total_drive_miles:,.0f}")

        self._cost_labor.setText(f"${labor_cost:,.2f}")
        if hasattr(self, "_cost_labor_formula"):
            _parts = []
            if regular_hrs > 0:
                _parts.append(f"{regular_hrs:.1f}h × ${_lr:,.0f}")
            if ot_hrs > 0:
                _parts.append(f"{ot_hrs:.1f}h × ${_lr * 1.5:,.0f} (OT)")
            if weekend_hrs > 0:
                _parts.append(f"{weekend_hrs:.1f}h × ${_lr * 2.0:,.0f} (wknd)")
            self._cost_labor_formula.setText(" + ".join(_parts) if _parts
                                              else f"{onsite_hrs:.1f} hrs × ${_lr:,.0f}/hr")
        self._cost_travel.setText(f"${travel_cost:,.2f}")
        if hasattr(self, "_cost_travel_formula"):
            self._cost_travel_formula.setText(
                f"{travel_hrs:.1f} hrs × ${_tr:,.0f}/hr" if travel_hrs else "—")
        self._cost_hotel.setText(f"${hotel_cost:,.2f}")
        if hasattr(self, "_cost_hotel_formula"):
            self._cost_hotel_formula.setText(
                f"{total_hotel} night{'s' if total_hotel != 1 else ''} × ${_hr:,.0f}/night")
        self._cost_meals.setText(f"${meal_cost:,.2f}")
        if hasattr(self, "_cost_meals_formula"):
            self._cost_meals_formula.setText(
                f"{total_meals} meal{'s' if total_meals != 1 else ''} × ${_mr:,.0f}/meal")
        self._cost_miles.setText(f"${mile_cost:,.2f}")
        if hasattr(self, "_cost_miles_formula"):
            self._cost_miles_formula.setText(
                f"{total_drive_miles:,.0f} mi × ${_mi:.3f}/mi" if total_drive_miles else "—")
        self._cost_flights.setText(f"${flight_total:,.2f}")
        if hasattr(self, "_cost_flights_formula"):
            self._cost_flights_formula.setText(
                f"{flights} flight{'s' if flights != 1 else ''}" if flights else "—")

        self._summary_card.setVisible(True)
        self._recalc_cost_totals()
        if hasattr(self, "_export_btn"):
            self._export_btn.setEnabled(True)
            self._export_btn.setToolTip("")

    # ── Cost helpers ──────────────────────────────────────────────────────────
    def _recalc_cost_totals(self):
        if not self._confirmed:
            return
        c = self._confirmed
        try:
            margin_pct = float(self._margin_lbl.text() or "0") / 100.0
        except Exception:
            margin_pct = 0.0
        subtotal = (c["labor_cost"] + c["travel_cost"] + c["hotel_cost"] +
                    c["meal_cost"]  + c["mile_cost"]   + c["flight_total"])
        grand = subtotal * (1.0 + margin_pct)
        self._cost_subtotal.setText(f"${subtotal:,.2f}")
        self._cost_grand.setText(f"${grand:,.2f}")

    def _on_meals_edited(self):
        """Recalculate meal cost when the user manually edits the meals count."""
        if not self._confirmed:
            return
        try:
            new_meals = int(float(self._sum_meals.text().replace(",", "").strip() or "0"))
        except (ValueError, TypeError):
            return
        new_meals = max(0, new_meals)
        self._sum_meals.setText(str(new_meals))
        new_meal_cost = new_meals * self.MEAL_RATE
        self._confirmed["meals"] = new_meals
        self._confirmed["meal_cost"] = new_meal_cost
        self._cost_meals.setText(f"${new_meal_cost:,.2f}")
        if hasattr(self, "_cost_meals_formula"):
            self._cost_meals_formula.setText(
                f"{new_meals} meal{'s' if new_meals != 1 else ''} × ${self.MEAL_RATE:,.0f}/meal")
        self._recalc_cost_totals()

    def _adj_margin(self, delta: int):
        try: v = float(self._margin_lbl.text() or "0")
        except: v = 0.0
        v = max(0.0, min(99.9, v + delta))
        self._margin_lbl.setText(f"{v:.1f}")

    def _update_wh_window_lbl(self):
        """Recompute and display the work-window duration next to the time dropdowns."""
        if not hasattr(self, "_work_start_combo") or not hasattr(self, "_wh_window_lbl"):
            return
        s = self._work_start_combo.currentIndex()
        e = self._work_end_combo.currentIndex()
        if e > s:
            hrs = e - s
            self._wh_window_lbl.setText(
                f"= {hrs}h/day  ·  weekday OT × 1.5  ·  weekend × 2.0")
        elif e < s:
            hrs = 24 - s + e
            self._wh_window_lbl.setText(
                f"= {hrs}h overnight (end is next day)  ·  weekday OT × 1.5  ·  weekend × 2.0")
        else:
            self._wh_window_lbl.setText("= 0h window — every hour counts as OT")

    def _mark_outlook_stale(self, *_):
        self._confirmed = {}
        if hasattr(self, "_confirm_btn"):
            self._confirm_btn.setEnabled(False)
            self._confirm_btn.setToolTip("Regenerate the Work Week Overview after changing config.")
        if hasattr(self, "_export_btn"):
            self._export_btn.setEnabled(False)
            self._export_btn.setToolTip("Confirm the schedule first.")

    def _update_gen_btn_state(self, *_):
        if hasattr(self, "_gen_btn"):
            has_panels = bool(self._panels_edit.text().strip())
            self._gen_btn.setEnabled(has_panels)
            self._gen_btn.setToolTip("" if has_panels else "Enter the total panel count first.")

    # ── Persistence ──────────────────────────────────────────────────────────
    def get_data(self) -> dict:
        # Capture live hotel checkbox states before serialising
        hotel_states = []
        if self._schedule and self._hotel_checks:
            for t in range(self._num_techs):
                tech_states = []
                for r in range(len(self._schedule)):
                    cb = (self._hotel_checks[t][r]
                          if t < len(self._hotel_checks)
                             and r < len(self._hotel_checks[t])
                          else None)
                    tech_states.append(bool(cb.isChecked()) if cb is not None else False)
                hotel_states.append(tech_states)

        return {
            "level":             self._level_combo.currentData(),
            "panels":            self._panels_edit.text(),
            "hours":             self._hours_edit.text(),
            "num_techs":         self._num_techs,
            "tech_names":        [r["name"].text()         for r in self._tech_rows],
            "tech_travel":       [r["travel"].text()        for r in self._tech_rows],
            "tech_modes":        [r["mode"].currentText()   for r in self._tech_rows],
            "tech_flight_costs": [r["flight_cost"].text()   for r in self._tech_rows],
            "tech_mileages":     [r["mileage"].text()       for r in self._tech_rows],
            "work_days":         [i for i, cb in enumerate(self._day_checks) if cb.isChecked()],
            "start_day":         (self._start_day_combo.currentIndex()
                                  if hasattr(self, "_start_day_combo") else 0),
            "work_start":        (self._work_start_combo.currentIndex()
                                  if hasattr(self, "_work_start_combo") else 8),
            "work_end":          (self._work_end_combo.currentIndex()
                                  if hasattr(self, "_work_end_combo")   else 17),
            "row_hours":         {str(r): (he.text() or "")
                                  for r, he in self._row_hours_edits.items()},
            # Schedule state (populated after Generate / Confirm)
            "schedule":      [(dl, act, insp, dh) for dl, act, insp, dh in self._schedule],
            "hotel_states":  hotel_states,
            "confirmed":     self._confirmed,
            "margin":        self._margin_lbl.text() if hasattr(self, "_margin_lbl") else "0.0",
        }

    def restore_data(self, d: dict):
        if not d:
            return
        lvl = d.get("level", 1)
        for i in range(self._level_combo.count()):
            if self._level_combo.itemData(i) == lvl:
                self._level_combo.setCurrentIndex(i)
                break
        self._panels_edit.setText(str(d.get("panels", "")))
        self._hours_edit.setText(str(d.get("hours", "8")))
        self._num_techs = d.get("num_techs", 1)
        self._tech_count_lbl.setText(str(self._num_techs))
        self._rebuild_tech_rows()
        for i, nm in enumerate(d.get("tech_names", [])):
            if i < len(self._tech_rows):
                self._tech_rows[i]["name"].setText(nm)
        for i, tv in enumerate(d.get("tech_travel", [])):
            if i < len(self._tech_rows):
                self._tech_rows[i]["travel"].setText(tv)
        for i, mode in enumerate(d.get("tech_modes", [])):
            if i < len(self._tech_rows):
                combo = self._tech_rows[i]["mode"]
                idx = combo.findText(mode)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
        for i, fc in enumerate(d.get("tech_flight_costs", [])):
            if i < len(self._tech_rows):
                self._tech_rows[i]["flight_cost"].setText(fc)
        for i, mi in enumerate(d.get("tech_mileages", [])):
            if i < len(self._tech_rows):
                self._tech_rows[i]["mileage"].setText(mi)
        wd = set(d.get("work_days", [0, 1, 2, 3, 4]))
        for i, cb in enumerate(self._day_checks):
            cb.setChecked(i in wd)
        if hasattr(self, "_start_day_combo"):
            self._start_day_combo.setCurrentIndex(int(d.get("start_day", 0)))
        if hasattr(self, "_work_start_combo"):
            self._work_start_combo.setCurrentIndex(int(d.get("work_start", 8)))
        if hasattr(self, "_work_end_combo"):
            self._work_end_combo.setCurrentIndex(int(d.get("work_end", 17)))
        self._recalc_preview()
        if hasattr(self, "_margin_lbl"):
            self._margin_lbl.setText(str(d.get("margin", "0.0")))


# ── IBECard — embeds IBEWidget inside a collapsible cost-estimator card ──────
class IBECard(_Card):
    """
    Collapsible card version of the IBE estimator for embedding in the
    Cost Estimator widget.  An 'Active' checkbox in the header enables/disables
    the section; when inactive the body stays hidden.
    """

    def __init__(self, on_change, parent=None):
        super().__init__("IBE Estimator", parent)
        self._on_change_cb = on_change

        # Active checkbox in card header
        self._enable_cb = QCheckBox("Active")
        self._enable_cb.setChecked(False)
        self._enable_cb.setStyleSheet(
            "QCheckBox { font-size:12px; color:#920d2e; font-weight:700;"
            "  border:none; padding:0; }"
            "QCheckBox::indicator { width:14px; height:14px; }"
            "QCheckBox::indicator:unchecked {"
            "  border:2px solid #d6c0c5; background:#ffffff; border-radius:3px; }"
            "QCheckBox::indicator:checked {"
            "  border:2px solid #920d2e; background:#920d2e; border-radius:3px; }"
        )
        self._enable_cb.toggled.connect(self._on_enable_changed)
        self._hdr_row.addWidget(self._enable_cb)

        # Compact IBEWidget (no scroll wrapper, no bottom bar)
        self._ibe = IBEWidget(compact=True)
        self._body_layout.addWidget(self._ibe)

        # Start inactive / collapsed
        self._body.setVisible(False)
        self._collapsed = True
        self._toggle_btn.setText("►")

    def _on_enable_changed(self, checked: bool):
        self._body.setVisible(checked)
        self._collapsed = not checked
        self._toggle_btn.setText("▼" if checked else "►")

    def get_data(self) -> dict:
        return {"active": self._enable_cb.isChecked(), **self._ibe.get_data()}

    def restore_data(self, d: dict):
        self._ibe.restore_data(d)
        active = d.get("active", False)
        self._enable_cb.blockSignals(True)
        self._enable_cb.setChecked(active)
        self._enable_cb.blockSignals(False)
        self._on_enable_changed(active)

    def subtotal(self) -> float:
        return 0.0


# ── Standalone IBE Excel export ───────────────────────────────────────────────

def _export_ibe_excel(data: dict, path: str,
                      labor_rate:  float = 125.0,
                      travel_rate: float = 100.0,
                      hotel_rate:  float = 150.0,
                      meal_rate:   float =  25.0,
                      mile_rate:   float =   0.72):
    """Export IBE work-week schedule to a formatted Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as _gcl

    wb = Workbook()

    # ── Palette ───────────────────────────────────────────────────────────────
    RED     = "920D2E"
    RED_LT  = "F5D0DA"
    GRAY    = "F4F6FA"
    BLUE_BG = "E8F0FE"
    BLUE_FG = "1A3A6E"
    WHITE   = "FFFFFF"
    DARK    = "1A0509"
    BDR_C   = "D6C0C5"

    thin = Side(style="thin", color=BDR_C)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fill(h): return PatternFill("solid", fgColor=h)
    def _font(bold=False, color=DARK, size=11, italic=False):
        return Font(name="Aptos", bold=bold, color=color, size=size, italic=italic)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _num_s(v):
        try: return float(str(v).replace(",", "").replace("$", ""))
        except: return 0.0

    # ── Tech palettes (matches cost_estimator) ────────────────────────────────
    _TECH_PALETTES = [
        ("920d2e", "FFF0F3"), ("1F4E79", "EBF3FB"),
        ("375623", "EBF5E8"), ("7B3F00", "FFF3E8"), ("4B0082", "F3EBF9"),
    ]
    def _thdr(ti): return _TECH_PALETTES[ti % len(_TECH_PALETTES)][0]
    def _tdat(ti): return _TECH_PALETTES[ti % len(_TECH_PALETTES)][1]

    # ── Data ──────────────────────────────────────────────────────────────────
    schedule     = data.get("schedule",      [])
    hotel_states = data.get("hotel_states",  [])
    n_techs      = data.get("num_techs",      1)
    tech_names   = data.get("tech_names",    [])
    tech_modes   = data.get("tech_modes",    [])
    tech_travel  = data.get("tech_travel",   [])
    tech_miles   = data.get("tech_mileages", [])
    tech_flight  = data.get("tech_flight_costs", [])
    confirmed    = data.get("confirmed",     {})
    ibe_lvl      = data.get("level",          1)
    ibe_rate_v   = {1: 7.5, 2: 5.0, 3: 3.375, 4: 1.875}.get(ibe_lvl, 7.5)

    # ── Sheet 1: Schedule ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "IBE Schedule"

    PER_T      = 7
    BASE       = 5
    TOTAL_COLS = 4 + n_techs * PER_T
    SUB_LABELS = ["Travel (hrs)", "Miles", "Mileage $", "Hotel?",
                  "Hotel Cost",   "Meals", "Meals Cost"]
    sub_widths = [13, 14, 11, 8, 11, 7, 11]

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 11
    for ti in range(n_techs):
        for ci in range(PER_T):
            ws.column_dimensions[_gcl(BASE + ti * PER_T + ci)].width = sub_widths[ci]

    sr = 1

    # Title
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=TOTAL_COLS)
    tc = ws.cell(sr, 1, "IBE Work Week Schedule")
    tc.font = _font(bold=True, color=WHITE, size=14)
    tc.fill = _fill(RED); tc.alignment = _align("center")
    ws.row_dimensions[sr].height = 26; sr += 1

    # Config line
    ws.cell(sr, 1, f"IBE Level {ibe_lvl}  —  {ibe_rate_v:g} panels/hr").font = _font(bold=True)
    _ws_h = int(data.get("work_start", 8))
    _we_h = int(data.get("work_end",   17))
    def _fmt_h(h):
        if h == 0:  return "12:00 AM"
        if h < 12:  return f"{h}:00 AM"
        if h == 12: return "12:00 PM"
        return f"{h - 12}:00 PM"
    if _we_h > _ws_h:   _wh_info = f"{_we_h - _ws_h}h/day"
    elif _we_h < _ws_h: _wh_info = f"{24 - _ws_h + _we_h}h overnight"
    else:               _wh_info = "all OT"
    ws.cell(sr, 2,
        f"{data.get('panels','?')} panels   "
        f"{n_techs} technician(s)   "
        f"Work Hours: {_fmt_h(_ws_h)} – {_fmt_h(_we_h)} ({_wh_info})   "
        f"OT: weekday × 1.5  ·  Weekend: × 2.0"
    ).font = _font()
    ws.row_dimensions[sr].height = 16; sr += 1

    # Rates
    ws.cell(sr, 1, "Rates:").font = _font(bold=True)
    ws.cell(sr, 2,
        f"Hotel ${hotel_rate:,.2f}/night     Meals ${meal_rate:,.2f}/meal"
        f"     Mileage ${mile_rate:,.3f}/mi"
    ).font = _font(italic=True, color="555555")
    ws.row_dimensions[sr].height = 16; sr += 2

    # Tech name header row
    for col, lbl in [(1,"Day"),(2,"Activity"),(3,"Panels Done"),(4,"% Done")]:
        c = ws.cell(sr, col, lbl)
        c.font = _font(bold=True); c.fill = _fill(GRAY); c.alignment = _align("center", wrap=True)
    for ti in range(n_techs):
        nm = tech_names[ti] if ti < len(tech_names) else f"Tech {ti+1}"
        sc = BASE + ti * PER_T
        ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + PER_T - 1)
        nc = ws.cell(sr, sc, nm)
        nc.font = _font(bold=True, color=WHITE)
        nc.fill = _fill(_thdr(ti)); nc.alignment = _align("center")
    ws.row_dimensions[sr].height = 22; sr += 1

    # Sub-header row
    for col in (1, 2, 3, 4):
        ws.cell(sr, col).fill = _fill(GRAY)
    for ti in range(n_techs):
        sc = BASE + ti * PER_T
        for ci, lbl in enumerate(SUB_LABELS):
            c = ws.cell(sr, sc + ci, lbl)
            c.font = Font(name="Aptos", bold=True, color=WHITE, size=10)
            c.fill = _fill(_thdr(ti)); c.alignment = _align("center", wrap=True)
    ws.row_dimensions[sr].height = 30; sr += 1

    # Row data
    total_panels_f = _num_s(data.get("panels", "0"))
    # Default hours from the work window, falling back to the Hours/Day field
    _ws_idx = int(data.get("work_start", 8))
    _we_idx = int(data.get("work_end",   17))
    if _we_idx > _ws_idx:   _win_hrs_f = float(_we_idx - _ws_idx)
    elif _we_idx < _ws_idx: _win_hrs_f = float(24 - _ws_idx + _we_idx)
    else:                   _win_hrs_f = _num_s(data.get("hours", "8"))
    hours_day_f    = _win_hrs_f if _win_hrs_f > 0 else _num_s(data.get("hours", "8"))
    row_hours_dict = data.get("row_hours", {})   # per-row overrides
    cum_panels     = 0.0

    def _blank_acc():
        return [{"t_hrs": 0.0, "miles": 0.0, "mile_cost": 0.0,
                 "hotel_nights": 0, "hotel_cost": 0.0,
                 "meals": 0, "meals_cost": 0.0} for _ in range(n_techs)]

    week_acc  = _blank_acc()
    grand_acc = _blank_acc()
    current_week     = 0
    prev_was_weekend = False

    def _write_week_banner(num):
        nonlocal sr
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=TOTAL_COLS)
        wc = ws.cell(sr, 1, f"WEEK {num}")
        wc.font = _font(bold=True, color=WHITE); wc.fill = _fill("5B6FA8")
        wc.alignment = _align("center")
        ws.row_dimensions[sr].height = 16; sr += 1

    def _write_subtotal(label, acc, fill_hex, txt_color=DARK):
        nonlocal sr
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=4)
        lc = ws.cell(sr, 1, label)
        lc.font = _font(bold=True, color=txt_color)
        lc.fill = _fill(fill_hex); lc.alignment = _align("center")
        for col in (2, 3, 4):
            ws.cell(sr, col).fill = _fill(fill_hex)
        for ti in range(n_techs):
            a  = acc[ti]; sc = BASE + ti * PER_T
            vals = [
                f"{a['t_hrs']:.1f} hrs"   if a['t_hrs']        else "—",
                f"{a['miles']:,.0f} mi"   if a['miles']        else "—",
                f"${a['mile_cost']:,.2f}" if a['mile_cost']    else "—",
                str(a['hotel_nights'])     if a['hotel_nights'] else "—",
                f"${a['hotel_cost']:,.2f}" if a['hotel_cost']  else "—",
                str(a['meals'])            if a['meals']        else "—",
                f"${a['meals_cost']:,.2f}" if a['meals_cost']  else "—",
            ]
            for ci, v in enumerate(vals):
                c = ws.cell(sr, sc + ci, v)
                c.font = _font(bold=True, color=txt_color)
                c.fill = _fill(fill_hex); c.alignment = _align("center")
        ws.row_dimensions[sr].height = 18; sr += 1

    for ri, row_data in enumerate(schedule):
        day_lbl, activity, is_insp, def_hotel = (list(row_data) + ["", "", False, False])[:4]
        is_weekend    = str(day_lbl) == "Weekend"
        is_travel_in  = activity == "Travel In"
        is_travel_out = activity == "Travel Out"
        is_travel     = is_travel_in or is_travel_out

        if is_weekend:
            _write_subtotal(f"Week {current_week} Subtotal", week_acc, "D9E1F2", "1F4E79")
            week_acc = _blank_acc()

        if ri == 0:
            current_week = 1; _write_week_banner(current_week)
        elif not is_weekend and prev_was_weekend:
            current_week += 1; _write_week_banner(current_week)

        prev_was_weekend = is_weekend
        row_bg = "F2F4F8" if is_weekend else WHITE

        dc = ws.cell(sr, 1, str(day_lbl))
        dc.font = _font(bold=not is_weekend, color=("666666" if is_weekend else DARK))
        dc.fill = _fill(row_bg); dc.alignment = _align("center")

        ac = ws.cell(sr, 2, str(activity))
        ac.font = _font(italic=is_weekend, color=("888888" if is_weekend else DARK))
        ac.fill = _fill(row_bg)

        if is_insp:
            _rh_txt = row_hours_dict.get(str(ri), "")
            try:    _row_hrs = float(_rh_txt) if _rh_txt else hours_day_f
            except: _row_hrs = hours_day_f
            cum_panels = min(cum_panels + ibe_rate_v * _row_hrs * n_techs, total_panels_f)
            pct = (cum_panels / total_panels_f * 100) if total_panels_f > 0 else 0
            pd_c  = ws.cell(sr, 3, f"{int(round(cum_panels)):,}")
            pct_c = ws.cell(sr, 4, f"{min(pct, 100.0):.1f}%")
            pd_c.font = pct_c.font = _font(bold=True, color=RED)
        else:
            pd_c  = ws.cell(sr, 3, "—"); pct_c = ws.cell(sr, 4, "—")
            pd_c.font = pct_c.font = _font(color="AAAAAA")
        pd_c.fill  = _fill(row_bg); pd_c.alignment  = _align("center")
        pct_c.fill = _fill(row_bg); pct_c.alignment = _align("center")

        for ti in range(n_techs):
            sc     = BASE + ti * PER_T
            mode   = tech_modes[ti]  if ti < len(tech_modes)  else "Driving"
            ow_mi  = _num_s(tech_miles[ti]  if ti < len(tech_miles)  else "0")
            f_cost = _num_s(tech_flight[ti] if ti < len(tech_flight) else "0") or 450.0
            t_hrs  = _num_s(tech_travel[ti] if ti < len(tech_travel) else "0")
            flying = (mode == "Flying")
            data_bg = _tdat(ti)

            hotel_booked = (
                bool(hotel_states[ti][ri])
                if ti < len(hotel_states) and ri < len(hotel_states[ti])
                else bool(def_hotel)
            )

            if is_weekend:
                rt_hrs  = t_hrs * 2
                rt_mi   = ow_mi * 2 if not flying else 0.0
                mi_cost = rt_mi * mile_rate if not flying else f_cost * 2
                mi_str  = (f"{rt_mi:,.0f} mi RT" if rt_mi else "—") if not flying else "Flying RT"
                vals = [
                    f"{rt_hrs:.1f} hrs" if rt_hrs else "—",
                    mi_str, f"${mi_cost:,.2f}" if mi_cost else "—",
                    "Home", "—", "—", "—",
                ]
                week_acc[ti]["t_hrs"]      += rt_hrs
                week_acc[ti]["miles"]      += rt_mi
                week_acc[ti]["mile_cost"]  += mi_cost
                grand_acc[ti]["t_hrs"]     += rt_hrs
                grand_acc[ti]["miles"]     += rt_mi
                grand_acc[ti]["mile_cost"] += mi_cost
                fills = ["F2F4F8"] * PER_T

            elif is_travel and not hotel_booked:
                # Local tech — no dedicated travel day; nothing to show
                vals  = ["—"] * PER_T
                fills = [data_bg] * PER_T

            else:
                if is_travel:
                    # Overnight tech travelling in/out
                    row_t  = t_hrs
                    row_mi = ow_mi if not flying else 0.0
                elif is_insp and not hotel_booked and not flying:
                    # Local commuter: round-trip each day
                    row_t  = t_hrs * 2
                    row_mi = ow_mi * 2
                else:
                    row_t  = 0.0; row_mi = 0.0

                mi_cost = row_mi * mile_rate
                n_meals = 2
                h_cost  = hotel_rate if hotel_booked else 0.0
                m_cost  = n_meals * meal_rate

                if flying and is_travel:
                    mi_str  = f"Flying  (${f_cost:,.0f}/trip)"
                    mi_cost = f_cost if is_travel_in else 0.0
                elif row_mi > 0:
                    sfx    = " RT" if (is_insp and not hotel_booked) else " OW"
                    mi_str = f"{row_mi:,.0f} mi{sfx}"
                else:
                    mi_str = "—"

                vals = [
                    f"{row_t:.1f} hrs" if row_t else "—",
                    mi_str, f"${mi_cost:,.2f}" if mi_cost else "—",
                    "Yes" if hotel_booked else "No",
                    f"${h_cost:,.2f}" if h_cost else "—",
                    str(n_meals), f"${m_cost:,.2f}",
                ]
                fills = [
                    data_bg, data_bg, data_bg,
                    ("FFF0F3" if hotel_booked else data_bg),
                    ("FFF0F3" if hotel_booked else data_bg),
                    data_bg, data_bg,
                ]
                week_acc[ti]["t_hrs"]        += row_t
                week_acc[ti]["miles"]        += row_mi
                week_acc[ti]["mile_cost"]    += mi_cost
                week_acc[ti]["hotel_nights"] += (1 if hotel_booked else 0)
                week_acc[ti]["hotel_cost"]   += h_cost
                week_acc[ti]["meals"]        += n_meals
                week_acc[ti]["meals_cost"]   += m_cost
                grand_acc[ti]["t_hrs"]        += row_t
                grand_acc[ti]["miles"]        += row_mi
                grand_acc[ti]["mile_cost"]    += mi_cost
                grand_acc[ti]["hotel_nights"] += (1 if hotel_booked else 0)
                grand_acc[ti]["hotel_cost"]   += h_cost
                grand_acc[ti]["meals"]        += n_meals
                grand_acc[ti]["meals_cost"]   += m_cost

            for ci, (v, sf) in enumerate(zip(vals, fills)):
                c = ws.cell(sr, sc + ci, v)
                c.font      = _font(color="666666" if is_weekend else DARK)
                c.fill      = _fill(sf)
                c.alignment = _align("center")

        ws.row_dimensions[sr].height = 18; sr += 1

    _write_subtotal(f"Week {current_week} Subtotal", week_acc, "D9E1F2", "1F4E79")
    sr += 1
    _write_subtotal("GRAND TOTAL", grand_acc, RED, WHITE)

    # Confirmed totals summary
    if confirmed:
        sr += 2
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=TOTAL_COLS)
        hdr2 = ws.cell(sr, 1, "Confirmed Schedule Totals")
        hdr2.font = _font(bold=True, color=WHITE); hdr2.fill = _fill(RED)
        hdr2.alignment = _align("center")
        ws.row_dimensions[sr].height = 20; sr += 1
        _reg_h = float(confirmed.get("regular_hrs",  0))
        _ot_h  = float(confirmed.get("ot_hrs",        0))
        _wk_h  = float(confirmed.get("weekend_hrs",   0))
        _tr_h  = float(confirmed.get("travel_hrs",    0))
        _tr_c  = float(confirmed.get("travel_cost",   0))
        _lb_c  = float(confirmed.get("labor_cost",    0))

        for lbl, val in [
            ("Inspection Days",             str(confirmed.get("insp_days",    "—"))),
            ("Hotel Nights",                str(confirmed.get("hotel_nights", "—"))),
            ("Total Meals",                 str(confirmed.get("meals",        "—"))),
            ("Drive Miles",                 f"{confirmed.get('drive_miles', 0):,.0f}"
                                             if confirmed.get("drive_miles") else "—"),
            ("Flights",                     str(confirmed.get("flights",      "—"))),
            ("Regular On-Site Hrs",         f"{_reg_h:.1f} hrs  ×  ${labor_rate:,.2f}/hr"
                                             f"  =  ${_reg_h * labor_rate:,.2f}"),
            ("OT On-Site Hrs (×1.5)",       f"{_ot_h:.1f} hrs  ×  ${labor_rate * 1.5:,.2f}/hr"
                                             f"  =  ${_ot_h * labor_rate * 1.5:,.2f}"),
            ("Weekend On-Site Hrs (×2.0)",  f"{_wk_h:.1f} hrs  ×  ${labor_rate * 2.0:,.2f}/hr"
                                             f"  =  ${_wk_h * labor_rate * 2.0:,.2f}"),
            ("On-Site Labor Total",         f"${_lb_c:,.2f}"),
            ("Travel Hrs",                  f"{_tr_h:.1f} hrs  ×  ${travel_rate:,.2f}/hr"
                                             f"  =  ${_tr_c:,.2f}"),
        ]:
            ws.cell(sr, 1, lbl).font = _font(bold=True)
            ws.cell(sr, 2, val).font = _font()
            ws.row_dimensions[sr].height = 16; sr += 1

    # ── Sheet 2: Cost Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Cost Summary")
    ws2.column_dimensions["A"].width = 52   # wider for formula strings
    ws2.column_dimensions["B"].width = 18

    cr = 1
    ws2.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=2)
    tc2 = ws2.cell(cr, 1, "IBE Schedule — Cost Summary")
    tc2.font = _font(bold=True, color=WHITE, size=14)
    tc2.fill = _fill(RED); tc2.alignment = _align("center")
    ws2.row_dimensions[cr].height = 26; cr += 2

    def _cs_row(label, value, bold=False, bg=WHITE, fg=DARK, h=17):
        ws2.cell(cr, 1, label).font = _font(bold=bold, color=fg)
        ws2.cell(cr, 1).fill = _fill(bg)
        c = ws2.cell(cr, 2, value)
        c.font = _font(bold=bold, color=fg); c.fill = _fill(bg)
        c.alignment = _align("right")
        ws2.row_dimensions[cr].height = h

    def _cs_section(title):
        """Bold section-header banner spanning both columns."""
        ws2.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=2)
        sh = ws2.cell(cr, 1, title)
        sh.font = _font(bold=True, color=WHITE, size=11)
        sh.fill = _fill(RED); sh.alignment = _align("left")
        ws2.row_dimensions[cr].height = 18

    # ── Labor section ─────────────────────────────────────────────────────────
    _reg_h  = confirmed.get("regular_hrs",  0.0)
    _ot_h   = confirmed.get("ot_hrs",        0.0)
    _wk_h   = confirmed.get("weekend_hrs",   0.0)
    _tr_h   = confirmed.get("travel_hrs",    0.0)
    _lab_c  = confirmed.get("labor_cost",    0.0)
    _tr_c   = confirmed.get("travel_cost",   0.0)

    _cs_section("LABOR — On-Site"); cr += 1

    _cs_row(
        f"  Regular  ({_reg_h:.1f} hrs  ×  ${labor_rate:,.2f}/hr)",
        f"${_reg_h * labor_rate:,.2f}", bg=WHITE
    ); cr += 1

    _cs_row(
        f"  Overtime ×1.5  ({_ot_h:.1f} hrs  ×  ${labor_rate * 1.5:,.2f}/hr)",
        f"${_ot_h * labor_rate * 1.5:,.2f}", bg=GRAY
    ); cr += 1

    _cs_row(
        f"  Weekend ×2.0  ({_wk_h:.1f} hrs  ×  ${labor_rate * 2.0:,.2f}/hr)",
        f"${_wk_h * labor_rate * 2.0:,.2f}", bg=WHITE
    ); cr += 1

    _cs_row(
        "  On-Site Labor Total",
        f"${_lab_c:,.2f}", bold=True, bg=RED_LT, fg="1A3A6E"
    ); cr += 2   # blank line

    _cs_section("LABOR — Travel"); cr += 1
    _cs_row(
        f"  Travel Time  ({_tr_h:.1f} hrs  ×  ${travel_rate:,.2f}/hr)",
        f"${_tr_c:,.2f}", bg=WHITE
    ); cr += 2   # blank line

    # ── Expenses section ──────────────────────────────────────────────────────
    _cs_section("EXPENSES"); cr += 1

    for lbl, key, shade in [
        ("  Hotel",   "hotel_cost",   False),
        ("  Meals",   "meal_cost",    True),
        ("  Mileage", "mile_cost",    False),
        ("  Flights", "flight_total", True),
    ]:
        v = confirmed.get(key, 0.0)
        _cs_row(lbl, f"${v:,.2f}", bg=GRAY if shade else WHITE)
        cr += 1

    # ── Totals ────────────────────────────────────────────────────────────────
    subtotal = sum(confirmed.get(k, 0.0) for k in
                   ("labor_cost", "travel_cost", "hotel_cost",
                    "meal_cost",  "mile_cost",   "flight_total"))
    try:
        margin_pct = float(data.get("margin", "0")) / 100.0
    except Exception:
        margin_pct = 0.0
    grand = subtotal * (1.0 + margin_pct)

    cr += 1
    _cs_row("Subtotal", f"${subtotal:,.2f}", bold=True, bg=RED_LT, fg="1A3A6E"); cr += 1
    if margin_pct:
        _cs_row(f"Margin ({margin_pct * 100:.1f}%)", f"{margin_pct * 100:.1f}%",
                bold=True, bg=BLUE_BG, fg=BLUE_FG); cr += 1
    _cs_row("Grand Total", f"${grand:,.2f}", bold=True, bg=RED, fg=WHITE, h=20); cr += 1

    wb.save(path)
